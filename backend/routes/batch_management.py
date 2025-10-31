from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from mongo import mongo_db
from bson import ObjectId
from datetime import datetime
import csv
import openpyxl
from werkzeug.utils import secure_filename
from config.constants import ROLES
from datetime import datetime
import pytz
import io
from utils.email_service import send_email, render_template
from utils.sms_service import send_student_credentials_sms
from utils.upload_optimizer import optimize_upload_process, cleanup_if_needed, log_upload_progress
from utils.resilient_services import create_resilient_services
from utils.async_processor import performance_monitor
from utils.notification_queue import queue_student_credentials, queue_batch_notifications, get_notification_stats
from config.shared import bcrypt
from socketio_instance import socketio
from routes.access_control import require_permission

batch_management_bp = Blueprint('batch_management', __name__)

def safe_isoformat(date_obj):
    """Safely convert a date object to ISO format string, handling various types."""
    if not date_obj:
        return None
    
    if hasattr(date_obj, 'isoformat'):
        # It's a datetime object
        return date_obj.isoformat()
    elif isinstance(date_obj, dict):
        # It's a MongoDB date dict, extract the date
        if '$date' in date_obj:
            try:
                from datetime import datetime
                date_str = date_obj['$date']
                # Handle different MongoDB date formats
                if 'T' in date_str:
                    # ISO format with T
                    date_obj_parsed = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                else:
                    # Just timestamp
                    date_obj_parsed = datetime.fromtimestamp(int(date_str) / 1000)
                return date_obj_parsed.isoformat()
            except (ValueError, KeyError, TypeError):
                return str(date_obj)
        else:
            return str(date_obj)
    else:
        # It's already a string or other type
        return str(date_obj)

@batch_management_bp.route('/', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_batches():
    """Get all batches"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Super admin and sub_superadmin can see all batches
        if user.get('role') in ['superadmin', 'sub_superadmin']:
            batches = list(mongo_db.batches.find())
        else:
            # Campus and course admins can only see batches in their campus
            campus_id = user.get('campus_id')
            if not campus_id:
                return jsonify({'success': False, 'message': 'No campus assigned'}), 400
            batches = list(mongo_db.batches.find({'campus_ids': ObjectId(campus_id)}))
        
        batch_list = []
        for batch in batches:
            # Get course details
            course_objs = list(mongo_db.courses.find({'_id': {'$in': batch.get('course_ids', [])}}))
            
            # Handle both old and new batch structures for campus data
            campus_ids = batch.get('campus_ids', [])
            if not campus_ids and batch.get('campus_id'):
                # Handle old structure with single campus_id
                campus_ids = [batch.get('campus_id')]
            
            # Get campus details
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': campus_ids}}))
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch.get('name'),
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'courses': [{'id': str(c['_id']), 'name': c['name']} for c in course_objs],
                'student_count': student_count,
                'created_at': batch.get('created_at')
            })
        
        return jsonify({'success': True, 'data': batch_list}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management', action='create_batch')
def create_batch_from_selection():
    """Create a new batch from selected campuses and courses"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Check if user has permission to create batches
        if not user or user.get('role') not in ['superadmin', 'sub_superadmin', 'campus_admin', 'course_admin']:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        data = request.get_json()
        name = data.get('name')
        campus_ids = [ObjectId(cid) for cid in data.get('campus_ids', [])]
        course_ids = [ObjectId(cid) for cid in data.get('course_ids', [])]

        if not name or not campus_ids or not course_ids:
            return jsonify({'success': False, 'message': 'Name, campuses, and courses are required'}), 400

        if mongo_db.batches.find_one({'name': name}):
            return jsonify({'success': False, 'message': 'Batch name already exists'}), 409

        # Campus and course admins can only create batches in their own campus
        if user.get('role') in ['campus_admin', 'course_admin']:
            user_campus_id = user.get('campus_id')
            if not user_campus_id or str(user_campus_id) not in [str(cid) for cid in campus_ids]:
                return jsonify({
                    'success': False,
                    'message': 'Access denied. You can only create batches in your own campus.'
                }), 403

        # Create the batch
        batch_id = mongo_db.batches.insert_one({
            'name': name,
            'campus_ids': campus_ids,
            'course_ids': course_ids,
            'created_at': datetime.now(pytz.utc)
        }).inserted_id

        # Create batch-course instances for each course
        created_instances = []
        for course_id in course_ids:
            instance_id = mongo_db.find_or_create_batch_course_instance(batch_id, course_id)
            created_instances.append({
                'batch_id': str(batch_id),
                'course_id': str(course_id),
                'instance_id': str(instance_id)
            })

        return jsonify({
            'success': True,
            'message': 'Batch created successfully',
            'data': {
                'batch_id': str(batch_id),
                'instances': created_instances
            }
        }), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/course/<course_id>/batches', methods=['GET'])
@jwt_required()
def get_batches_for_course(course_id):
    try:
        batches = list(mongo_db.batches.find({'course_ids': ObjectId(course_id)}))
        batch_list = []
        for batch in batches:
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': batch.get('campus_ids', [])}}))
            # Get student count for this batch
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch['name'],
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'student_count': student_count
            })
        return jsonify({'success': True, 'data': batch_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batches for course: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/campus/<campus_id>/batches', methods=['GET'])
@jwt_required()
def get_batches_for_campus(campus_id):
    try:
        batches = list(mongo_db.batches.find({'campus_ids': ObjectId(campus_id)}))
        batch_list = []
        for batch in batches:
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': batch.get('campus_ids', [])}}))
            course_objs = list(mongo_db.courses.find({'_id': {'$in': batch.get('course_ids', [])}}))
            # Get student count for this batch
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch['name'],
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'courses': [{'id': str(c['_id']), 'name': c['name']} for c in course_objs],
                'student_count': student_count
            })
        return jsonify({'success': True, 'data': batch_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batches for campus: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/create', methods=['POST'])
@jwt_required()
def create_batch():
    """Create a new batch and upload student data from an Excel file - SUPER ADMIN ONLY"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Only super admin and sub_superadmin can create batches
        if not user or user.get('role') not in ['superadmin', 'sub_superadmin']:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        if 'student_file' not in request.files:
            return jsonify({'success': False, 'message': 'No student file part'}), 400

        file = request.files['student_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No selected file'}), 400

        batch_name = request.form.get('batch_name')
        campus_id = request.form.get('campus_id')
        course_id = request.form.get('course_id')

        if not all([batch_name, campus_id, course_id]):
            return jsonify({'success': False, 'message': 'Missing batch name, campus ID, or course ID'}), 400

        # Create a new batch
        batch_doc = {
            'name': batch_name,
            'campus_id': ObjectId(campus_id),
            'course_id': ObjectId(course_id),
            'campus_ids': [ObjectId(campus_id)],
            'course_ids': [ObjectId(course_id)],
            'created_at': datetime.now(pytz.utc)
        }
        new_batch_id = mongo_db.batches.insert_one(batch_doc).inserted_id

        # Create batch-course instance
        instance_id = mongo_db.find_or_create_batch_course_instance(new_batch_id, ObjectId(course_id))

        # Process student file
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Get data rows
        students_data = []
        for row in worksheet.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(cell.value).strip() if cell.value else ''
            students_data.append(row_data)

        created_students = []
        for student in students_data:
            # Generate username and password
            username = str(student['roll_number'])
            password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
            
            # Use bcrypt for hashing
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

            student_doc = {
                'name': student['student_name'],
                'email': student['email_id'],
                'roll_number': str(student['roll_number']),
                'username': username,
                'password_hash': password_hash,
                'role': ROLES['STUDENT'],
                'campus_id': ObjectId(campus_id),
                'course_id': ObjectId(course_id),
                'batch_id': new_batch_id,
                'batch_course_instance_id': instance_id,  # Link to instance
                'is_active': True,
                'created_at': datetime.now(pytz.utc),
                'mfa_enabled': False
            }
            mongo_db.users.insert_one(student_doc)
            
            # Create student profile with instance link
            student_profile = {
                'user_id': student_doc['_id'],
                'name': student['student_name'],
                'roll_number': str(student['roll_number']),
                'email': student['email_id'],
                'campus_id': ObjectId(campus_id),
                'course_id': ObjectId(course_id),
                'batch_id': new_batch_id,
                'batch_course_instance_id': instance_id,  # Link to instance
                'created_at': datetime.now(pytz.utc)
            }
            mongo_db.students.insert_one(student_profile)
            
            created_students.append({
                "student_name": student['student_name'],
                "email_id": student['email_id'],
                "roll_number": str(student['roll_number']),
                "username": username,
                "password": password
            })

            # Send welcome email
            try:
                html_content = render_template(
                    'student_credentials.html',
                    params={
                        'name': student['student_name'],
                        'username': username,
                        'email': student['email_id'],
                        'password': password,
                        'login_url': "https://crt.pydahsoft.in/login"
                    }
                )
                send_email(
                    to_email=student['email_id'],
                    to_name=student['student_name'],
                    subject="Welcome to VERSANT - Your Student Credentials",
                    html_content=html_content
                )
            except Exception as e:
                print(f"Failed to send welcome email to {student['email_id']}: {e}")

        # TODO: Add push notification for batch creation

        return jsonify({
            'success': True,
            'message': 'Batch and students created successfully',
            'data': {
                'batch_id': str(new_batch_id),
                'instance_id': str(instance_id),
                'created_students': created_students
            }
        }), 201
    except Exception as e:
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>', methods=['PUT'])
@jwt_required()
def edit_batch(batch_id):
    try:
        data = request.get_json()
        name = data.get('name')

        if not name:
            return jsonify({'success': False, 'message': 'Batch name is required.'}), 400

        result = mongo_db.batches.update_one(
            {'_id': ObjectId(batch_id)},
            {'$set': {'name': name}}
        )

        if result.matched_count == 0:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404

        return jsonify({'success': True, 'message': 'Batch updated successfully.'}), 200
    except Exception as e:
        current_app.logger.error(f"Error updating batch: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while updating the batch.'}), 500

@batch_management_bp.route('/<batch_id>', methods=['DELETE'])
@jwt_required()
def delete_batch(batch_id):
    try:
        batch_obj_id = ObjectId(batch_id)
        
        # Find all students in this batch to get their user_ids
        students_to_delete = list(mongo_db.students.find({'batch_id': batch_obj_id}))
        user_ids_to_delete = [s['user_id'] for s in students_to_delete]
        
        # Delete associated users
        if user_ids_to_delete:
            mongo_db.users.delete_many({'_id': {'$in': user_ids_to_delete}})
        
        # Delete student records
        mongo_db.students.delete_many({'batch_id': batch_obj_id})
        
        # Finally, delete the batch
        result = mongo_db.batches.delete_one({'_id': batch_obj_id})
        
        if result.deleted_count == 0:
            return jsonify({'success': False, 'message': 'Batch not found or already deleted.'}), 404
            
        return jsonify({'success': True, 'message': 'Batch and all associated students have been deleted.'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error deleting batch: {e}")
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'}), 500

@batch_management_bp.route('/campuses', methods=['GET'])
@jwt_required()
def get_campuses():
    campuses = list(mongo_db.campuses.find())
    return jsonify({'success': True, 'data': [{'id': str(c['_id']), 'name': c['name']} for c in campuses]}), 200

@batch_management_bp.route('/courses', methods=['GET'])
@jwt_required()
def get_courses():
    campus_ids = request.args.getlist('campus_ids')
    if not campus_ids:
        return jsonify({'success': False, 'message': 'campus_ids required'}), 400
    try:
        campus_obj_ids = [ObjectId(cid) for cid in campus_ids]
        
        pipeline = [
            {'$match': {'campus_id': {'$in': campus_obj_ids}}},
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'campus_id',
                    'foreignField': '_id',
                    'as': 'campus_info'
                }
            },
            {
                '$unwind': '$campus_info'
            }
        ]
        courses = list(mongo_db.courses.aggregate(pipeline))

        courses_data = [{
            'id': str(c['_id']),
            'name': c['name'],
            'campus_id': str(c['campus_id']),
            'campus_name': c['campus_info']['name']
        } for c in courses]
        
        return jsonify({'success': True, 'data': courses_data}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching courses by campus: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

def _parse_student_file(file):
    filename = secure_filename(file.filename)
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        raise ValueError('Please upload a valid CSV or Excel file.')
    
    try:
        if filename.endswith('.csv'):
            # Read CSV file
            content = file.read().decode('utf-8-sig')
            csv_reader = csv.DictReader(io.StringIO(content))
            rows = list(csv_reader)
            
            # Filter out completely empty rows
            rows = [row for row in rows if any(str(value).strip() for value in row.values() if value is not None)]
            
            print(f"CSV parsing: Found {len(rows)} rows after filtering empty rows, headers: {list(rows[0].keys()) if rows else []}")
        else:
            # Read Excel file
            workbook = openpyxl.load_workbook(file, data_only=True)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            # Get data rows
            rows = []
            for row in worksheet.iter_rows(min_row=2):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = str(cell.value).strip() if cell.value else ''
                rows.append(row_data)
            
            # Filter out completely empty rows
            rows = [row for row in rows if any(str(value).strip() for value in row.values() if value is not None)]
            
            print(f"Excel parsing: Found {len(rows)} rows after filtering empty rows, headers: {headers}")
    except Exception as e:
        print(f"File parsing error: {e}")
        raise ValueError(f"Error reading file: {e}")

    return rows

@batch_management_bp.route('/test-file-parse', methods=['POST'])
@jwt_required()
def test_file_parse():
    """Test endpoint to debug file parsing"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'success': False, 'message': 'No file provided'}), 400

        rows = _parse_student_file(file)
        
        return jsonify({
            'success': True,
            'data': {
                'row_count': len(rows),
                'columns': list(rows[0].keys()) if rows else [],
                'first_row': rows[0] if rows else None,
                'sample_rows': rows[:3] if len(rows) > 3 else rows
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/validate-student-upload', methods=['POST'])
@jwt_required()
def validate_student_upload():
    try:
        file = request.files.get('file')
        campus_id = request.form.get('campus_id')

        if not file or not campus_id:
            return jsonify({'success': False, 'message': 'A file and campus ID are required.'}), 400

        rows = _parse_student_file(file)

        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400

        # Get column names from first row
        columns = list(rows[0].keys()) if rows else []
        
        required_fields = ['Campus Name', 'Course Name', 'Student Name', 'Roll Number', 'Email']
        missing_fields = [field for field in required_fields if field not in columns]
        if missing_fields:
            return jsonify({'success': False, 'message': f"Invalid file structure. Missing columns: {', '.join(missing_fields)}"}), 400

        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        # Note: Email duplicates are now allowed, so we don't need to check existing emails
        existing_emails = set()  # Empty set since we allow email duplicates
        existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))
        
        # Get campus info for validation
        campus = mongo_db.campuses.find_one({'_id': ObjectId(campus_id)})
        if not campus:
            return jsonify({'success': False, 'message': 'Invalid campus ID'}), 400
        
        campus_courses = list(mongo_db.courses.find({'campus_id': ObjectId(campus_id)}, {'name': 1}))
        valid_course_names = {course['name'] for course in campus_courses}

        preview_data = []
        for index, row in enumerate(rows):
            student_data = {
                'campus_name': str(row.get('Campus Name', '')).strip(),
                'course_name': str(row.get('Course Name', '')).strip(),
                'student_name': str(row.get('Student Name', '')).strip(),
                'roll_number': str(row.get('Roll Number', '')).strip(),
                'email': str(row.get('Email', '')).strip().lower(),
                'mobile_number': str(row.get('Mobile Number', '')).strip(),
            }
            
            errors = []
            if not all([student_data['campus_name'], student_data['course_name'], student_data['student_name'], student_data['roll_number'], student_data['email']]):
                errors.append('Missing required fields.')
            if student_data['roll_number'] in existing_roll_numbers:
                errors.append('Roll number already exists.')
            if student_data['email'] in existing_emails:
                errors.append('Email already exists.')
            if student_data['mobile_number'] and student_data['mobile_number'] in existing_mobile_numbers:
                errors.append('Mobile number already exists.')
            if student_data['campus_name'] != campus['name']:
                errors.append(f"Campus '{student_data['campus_name']}' doesn't match selected campus '{campus['name']}'.")
            if student_data['course_name'] not in valid_course_names:
                errors.append(f"Course '{student_data['course_name']}' not found in this campus.")

            student_data['errors'] = errors
            preview_data.append(student_data)
        
        return jsonify({'success': True, 'data': preview_data})

    except ValueError as ve:
        return jsonify({'success': False, 'message': str(ve)}), 400
    except Exception as e:
        current_app.logger.error(f"Error validating student upload: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {e}'}), 500

@batch_management_bp.route('/upload-students', methods=['POST'])
@jwt_required()
@performance_monitor(threshold=5.0)
def upload_students_to_batch():
    try:
        # Accept file upload and form data
        file = request.files.get('file')
        batch_id = request.form.get('batch_id')
        course_ids = request.form.getlist('course_ids')  # Accept multiple course IDs
        user_id = get_jwt_identity()  # Get current user ID for progress updates
        
        # Debug logging
        current_app.logger.info(f"Upload request - batch_id: {batch_id}, course_ids: {course_ids}, filename: {file.filename if file else 'None'}")
        
        if not file or not batch_id or not course_ids:
            return jsonify({'success': False, 'message': 'File, batch ID, and at least one course ID are required.'}), 400

        rows = _parse_student_file(file)
        current_app.logger.info(f"Parsed {len(rows)} rows from file")
        
        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400
        
        # Log first few rows for debugging
        current_app.logger.info(f"First 3 rows: {rows[:3] if len(rows) >= 3 else rows}")

        # Validate columns - support both formats (email is now optional)
        columns = list(rows[0].keys()) if rows else []
        current_app.logger.info(f"File columns: {columns}")
        
        required_fields_v1 = ['Student Name', 'Roll Number', 'Mobile Number']  # Removed Email from required
        required_fields_v2 = ['Group', 'Roll Number', 'Student Name', 'Mobile Number']  # Removed Email from required
        
        missing_fields_v1 = [field for field in required_fields_v1 if field not in columns]
        missing_fields_v2 = [field for field in required_fields_v2 if field not in columns]
        
        current_app.logger.info(f"Missing v1 fields: {missing_fields_v1}, Missing v2 fields: {missing_fields_v2}")
        
        if missing_fields_v1 and missing_fields_v2:
            return jsonify({'success': False, 'message': f"Invalid file structure. Expected either: {', '.join(required_fields_v1)} OR {', '.join(required_fields_v2)}. Found columns: {', '.join(columns)}"}), 400

        # Determine file format
        is_v2_format = 'Group' in columns
        current_app.logger.info(f"Using format v2: {is_v2_format}")

        # Fetch batch and campus info
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        campus_ids = batch.get('campus_ids', [])
        if not campus_ids:
            return jsonify({'success': False, 'message': 'Batch is missing campus info.'}), 400
        campus_id = campus_ids[0]  # Assume single campus per batch for now

        # Validate course IDs
        valid_course_ids = set(str(cid) for cid in batch.get('course_ids', []))
        for cid in course_ids:
            if cid not in valid_course_ids:
                return jsonify({'success': False, 'message': f'Course ID {cid} is not valid for this batch.'}), 400

        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        # Note: Email duplicates are now allowed, so we don't need to check existing emails
        existing_emails = set()  # Empty set since we allow email duplicates
        existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))

        # Detailed response tracking
        detailed_results = []
        errors = []
        total_students = len(rows)
        
        # Send initial progress update
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'started',
            'total': total_students,
            'processed': 0,
            'percentage': 0,
            'message': 'Starting student upload...'
        }, room=str(user_id))
        
        current_app.logger.info(f"Starting PHASED processing of {len(rows)} rows")
        
        # Optimize upload process for this batch size
        optimize_upload_process(len(rows))
        
        # Initialize all student results first
        detailed_results = []
        for index, row in enumerate(rows):
            student_result = {
                'student_name': '',
                'roll_number': '',
                'email': '',
                'mobile_number': '',
                'database_registered': False,
                'email_sent': False,
                'sms_sent': False,
                'errors': [],
                'success': False
            }
            detailed_results.append(student_result)
        
        # PHASE 1: DATABASE REGISTRATION
        current_app.logger.info("🚀 PHASE 1: Starting database registration...")
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'processing',
            'total': total_students,
            'processed': 0,
            'percentage': 0,
            'message': 'Phase 1: Registering students in database...'
        }, room=str(user_id))
        
        for index, row in enumerate(rows):
            student_result = detailed_results[index]
            
            # Log every 10th row for debugging
            if index % 10 == 0:
                current_app.logger.info(f"Database Phase - Processing row {index + 1}/{len(rows)}: {row}")
                # Perform cleanup and log progress
                cleanup_if_needed(index + 1)
                log_upload_progress(index + 1, len(rows), "Database Phase - ")
            
            try:
                if is_v2_format:
                    student_name = str(row.get('Student Name', '')).strip()
                    roll_number = str(row.get('Roll Number', '')).strip()
                    email = str(row.get('Email', '')).strip().lower() if row.get('Email') else ''  # Make email optional
                    mobile_number = str(row.get('Mobile Number', '')).strip()
                    group_name = str(row.get('Group', '')).strip()
                    
                    # Find course by group name (assuming group name matches course name)
                    # Try exact match first, then case-insensitive match
                    course = mongo_db.courses.find_one({'name': group_name, '_id': {'$in': [ObjectId(cid) for cid in course_ids]}})
                    if not course:
                        # Try case-insensitive match
                        course = mongo_db.courses.find_one({
                            'name': {'$regex': f'^{group_name}$', '$options': 'i'}, 
                            '_id': {'$in': [ObjectId(cid) for cid in course_ids]}
                        })
                    
                    if not course:
                        # Get available courses for better error message
                        available_courses = list(mongo_db.courses.find({'_id': {'$in': [ObjectId(cid) for cid in course_ids]}}, {'name': 1}))
                        available_names = [c['name'] for c in available_courses]
                        student_result['errors'].append(f"Course/Group '{group_name}' not found in this batch. Available courses: {', '.join(available_names)}")
                        detailed_results.append(student_result)
                        continue
                    course_id = str(course['_id'])
                else:
                    student_name = str(row.get('Student Name', '')).strip()
                    roll_number = str(row.get('Roll Number', '')).strip()
                    email = str(row.get('Email', '')).strip().lower() if row.get('Email') else ''  # Make email optional
                    mobile_number = str(row.get('Mobile Number', '')).strip()
                    course_id = course_ids[0]  # Use first course for v1 format

                # Update student result with basic info
                student_result.update({
                    'student_name': student_name,
                    'roll_number': roll_number,
                    'email': email,
                    'mobile_number': mobile_number
                })

            # Validation
                validation_errors = []
                
                # Check for missing required fields (name and roll number)
                if not student_name or not roll_number or student_name.strip() == '' or roll_number.strip() == '':
                    validation_errors.append('Missing required fields.')
                
                # Skip processing if missing required fields
                if validation_errors:
                    student_result['errors'] = validation_errors
                    detailed_results.append(student_result)
                    continue
                
                # Check for duplicates only if we have valid data
                if roll_number in existing_roll_numbers:
                    validation_errors.append('Roll number already exists.')
                # Note: Email duplicates are now allowed at database level
                # We can log a warning but don't treat it as an error
                if email and email.strip() and email in existing_emails:
                    current_app.logger.warning(f"⚠️ Email {email} already exists for another user, but allowing duplicate")
                if mobile_number and mobile_number in existing_mobile_numbers:
                    validation_errors.append('Mobile number already exists.')
                
                if validation_errors:
                    student_result['errors'] = validation_errors
                    detailed_results.append(student_result)
                    continue

                # Database Registration
                try:
                    username = roll_number
                    password = f"{student_name.split()[0][:4].lower()}{roll_number[-4:]}"
                    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                    
                    user_doc = {
                        'username': username,
                        'email': email if email and email.strip() else None,  # Always include email field
                        'password_hash': password_hash,
                        'role': 'student',
                        'name': student_name,
                        'mobile_number': mobile_number,
                        'campus_id': campus_id,
                        'course_id': ObjectId(course_id),
                        'batch_id': ObjectId(batch_id),
                        'is_active': True,
                        'created_at': datetime.now(pytz.utc),
                        'mfa_enabled': False
                    }
                    
                    # Insert user first
                    user_result = mongo_db.users.insert_one(user_doc)
                    if not user_result.inserted_id:
                        student_result['errors'].append('Failed to create user account.')
                        detailed_results.append(student_result)
                        continue
                    
                    user_id_inserted = user_result.inserted_id
                    
                    # Create default notification preferences for the new user
                    try:
                        from models_notification_preferences import NotificationPreferences
                        NotificationPreferences.create_default_preferences(user_id_inserted)
                        current_app.logger.info(f"✅ Created default notification preferences for student: {user_id_inserted}")
                    except Exception as e:
                        current_app.logger.warning(f"⚠️ Failed to create default notification preferences for student {user_id_inserted}: {e}")
                        # Don't fail user creation if notification preferences creation fails
                    
                    student_doc = {
                        'user_id': user_id_inserted,
                        'name': student_name,
                        'roll_number': roll_number,
                        'email': email if email and email.strip() else None,  # Always include email field
                        'mobile_number': mobile_number,
                        'campus_id': campus_id,
                        'course_id': ObjectId(course_id),
                        'batch_id': ObjectId(batch_id),
                        'created_at': datetime.now(pytz.utc)
                    }
                    
                    # Insert student profile
                    student_result_db = mongo_db.students.insert_one(student_doc)
                    if not student_result_db.inserted_id:
                        # Rollback user creation if student creation fails
                        mongo_db.users.delete_one({'_id': user_id_inserted})
                        student_result['errors'].append('Failed to create student profile.')
                        detailed_results.append(student_result)
                        continue
                    
                    # Database registration successful
                    student_result['database_registered'] = True
                    student_result['username'] = username
                    student_result['password'] = password
                    
                    # Update existing sets to prevent duplicates within the same upload
                    existing_roll_numbers.add(roll_number)
                    existing_emails.add(email)
                    if mobile_number:
                        existing_mobile_numbers.add(mobile_number)
                
                except Exception as e:
                    student_result['errors'].append(f'Database error: {str(e)}')
                    current_app.logger.error(f"Database error for student {index + 1}: {e}")
                    continue

                # Send progress update for database phase
                percentage = int(((index + 1) / total_students) * 100)  # Database phase is 100% of upload
                socketio.emit('upload_progress', {
                    'user_id': user_id,
                    'status': 'processing',
                    'total': total_students,
                    'processed': index + 1,
                    'percentage': percentage,
                    'message': f'Uploading: {student_name} - {"✅" if student_result["database_registered"] else "❌"}',
                    'current_student': {
                        'name': student_name,
                        'email': email,
                        'username': username if student_result['database_registered'] else '',
                        'database_registered': student_result['database_registered'],
                        'email_sent': False,
                        'sms_sent': False
                    }
                }, room=str(user_id))
                    
            except Exception as e:
                student_result['errors'].append(f'Database error: {str(e)}')
                current_app.logger.error(f"Database error for student {index + 1}: {e}")
                continue

        # Database upload completed - now queue notifications in background
        current_app.logger.info("✅ Database upload phase completed successfully!")

        # PHASE 2: QUEUE NOTIFICATIONS IN BACKGROUND
        current_app.logger.info("🚀 PHASE 2: Queueing notifications in background...")
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'queueing_notifications',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
            'message': 'Database upload completed! Queueing notifications in background...'
        }, room=str(user_id))

        # Prepare students for notification queueing
        students_for_notifications = []
        for index, (row, student_result) in enumerate(zip(rows, detailed_results)):
            if student_result['database_registered']:
                # Extract student data for notifications
                if is_v2_format:
                    student_name = str(row.get('Student Name', '')).strip()
                    roll_number = str(row.get('Roll Number', '')).strip()
                    email = str(row.get('Email', '')).strip().lower() if row.get('Email') else ''
                    mobile_number = str(row.get('Mobile Number', '')).strip()
                else:
                    student_name = str(row.get('Student Name', '')).strip()
                    roll_number = str(row.get('Roll Number', '')).strip()
                    email = str(row.get('Email', '')).strip().lower() if row.get('Email') else ''
                    mobile_number = str(row.get('Mobile Number', '')).strip()

                # Generate username and password (same logic as database registration)
                username = roll_number
                password = f"{student_name.split()[0][:4].lower()}{roll_number[-4:]}"

                # Get user_id from the student_result or find it by username
                user_id_for_student = None
                try:
                    # Try to find the user_id by username
                    user_doc = mongo_db.users.find_one({'username': username})
                    if user_doc:
                        user_id_for_student = str(user_doc['_id'])
                except Exception as e:
                    current_app.logger.warning(f"Could not find user_id for username {username}: {e}")

                student_data = {
                    'name': student_name,
                    'username': username,
                    'password': password,
                    'email': email if email and email.strip() else None,
                    'mobile_number': mobile_number if mobile_number and mobile_number.strip() else None,
                    'roll_number': roll_number,
                    'user_id': user_id_for_student  # Add user_id for push notifications
                }
                students_for_notifications.append(student_data)

        # Create batch job for credentials notifications (same as test creation)
        if students_for_notifications:
            try:
                from utils.batch_processor import create_credentials_batch_job
                
                # Create batch job for student credentials notifications
                batch_result = create_credentials_batch_job(
                    students=students_for_notifications,
                    batch_size=100,
                    interval_minutes=3
                )
                
                current_app.logger.info(f"📧📱 Student credentials batch created: {batch_result}")
                
                # Update student results with notification status
                for student_result in detailed_results:
                    if student_result['database_registered']:
                        student_result['notifications_queued'] = True
                        student_result['email_queued'] = True  # Will be processed in background
                        student_result['sms_queued'] = True    # Will be processed in background
                        student_result['push_queued'] = True   # Will be processed in background
                
                # Send ALL students to notification service at once (fire-and-forget)
                try:
                    import requests
                    import os
                    notification_service_url = os.getenv('NOTIFICATION_SERVICE_URL', 'http://localhost:3001')
                    notification_service_url = notification_service_url.rstrip('/api').rstrip('/')
                    
                    current_app.logger.info(f"📧📱 Sending {len(students_for_notifications)} students to notification service (batch)")
                    
                    # Send email batch (fire-and-forget)
                    try:
                        requests.post(
                            f"{notification_service_url}/api/email/send-credentials-batch",
                            json={'students': students_for_notifications, 'loginUrl': 'https://crt.pydahsoft.in/login'},
                            timeout=1  # Fire-and-forget
                        )
                        current_app.logger.info(f"✅ Email batch queued in notification service")
                    except:
                        pass  # Ignore errors
                    
                    # Send SMS batch (fire-and-forget)
                    try:
                        requests.post(
                            f"{notification_service_url}/api/sms/send-credentials-batch",
                            json={'students': students_for_notifications},
                            timeout=1  # Fire-and-forget
                        )
                        current_app.logger.info(f"✅ SMS batch queued in notification service")
                    except:
                        pass  # Ignore errors
                        
                except Exception as e:
                    current_app.logger.warning(f"⚠️ Failed to queue student credentials in notification service: {e}")
                    # Don't fail the upload if notifications fail
                    
            except Exception as e:
                current_app.logger.error(f"❌ Failed to create credentials batch: {e}")
                # Don't fail the upload if notifications fail
        else:
            current_app.logger.warning("⚠️ No students available for notification batch creation")

        # Finalize all student results
        for student_result in detailed_results:
            student_result['success'] = student_result['database_registered']
            # Set notification flags based on queueing success
            student_result['email_sent'] = student_result.get('email_queued', False)
            student_result['sms_sent'] = student_result.get('sms_queued', False)
            student_result['push_sent'] = student_result.get('push_queued', False)
        
        # Send completion notification
        current_app.logger.info("🎉 Student upload completed successfully with batch processing!")
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'completed',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
            'message': f'Upload completed! {len(students_for_notifications)} students queued for batch processing.'
        }, room=str(user_id))
                    
        # Calculate summary statistics (only database registrations)
        successful_registrations = sum(1 for r in detailed_results if r['database_registered'])
        total_errors = sum(len(r['errors']) for r in detailed_results)
        
        # Calculate additional statistics (database only)
        database_only = successful_registrations  # All successful registrations are database-only
        complete_failures = total_students - successful_registrations
        
        # Create summary for database upload only
        upload_summary = {
            'total_students': total_students,
            'database_registered': successful_registrations,
            'emails_queued': len(students_for_notifications),  # Emails queued for batch processing
            'sms_queued': len(students_for_notifications),     # SMS queued for batch processing
            'total_errors': total_errors,
            'database_only': database_only,
            'complete_failures': complete_failures,
            'success_rate': round((successful_registrations / total_students * 100), 2) if total_students > 0 else 0,
            'email_success_rate': 0,  # No emails sent during upload
            'sms_success_rate': 0     # No SMS sent during upload
        }

        # Send completion progress update
        if total_errors > 0:
            socketio.emit('upload_progress', {
                'user_id': user_id,
                'status': 'completed_with_errors',
                'total': total_students,
                'processed': total_students,
                'percentage': 100,
                'message': f'Database upload completed with {total_errors} errors. {successful_registrations} students registered.',
                'summary': upload_summary
            }, room=str(user_id))
            
            return jsonify({
                'success': successful_registrations > 0, 
                'message': f"Database upload completed with {total_errors} errors. {successful_registrations} students registered. Credentials notifications queued for batch processing.", 
                'data': {
                    'detailed_results': detailed_results,
                    'summary': upload_summary,
                    'status_breakdown': {
                        'database_only': database_only,
                        'complete_failures': complete_failures
                    }
                }
            }), 207 if total_errors > 0 else 201
        
        # Log final results for debugging
        current_app.logger.info(f"Final processing results: {len(detailed_results)} detailed results, {total_students} total students")
        current_app.logger.info(f"Summary: DB={successful_registrations}, Errors={total_errors}")
        
        # Send success completion update
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'completed',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
            'message': f'Successfully uploaded {total_students} students to database! Credentials notifications queued for batch processing.',
            'summary': upload_summary
        }, room=str(user_id))
        
        return jsonify({
            'success': True, 
            'message': f"Successfully uploaded {total_students} students to database. Credentials notifications queued for batch processing.", 
            'data': {
                'detailed_results': detailed_results,
                'summary': upload_summary,
                'status_breakdown': {
                    'database_only': database_only,
                    'complete_failures': complete_failures
                }
            }
        }), 201

    except Exception as e:
        current_app.logger.error(f"Error uploading students to batch: {e}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Failed to upload students. Please check your file format and try again. Error: {str(e)}'}), 500

@batch_management_bp.route('/batch/<batch_id>/students', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_batch_students(batch_id):
    """Get all students and detailed info for a specific batch."""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found'}), 404

        # Check if user has access to this batch
        if user.get('role') in ['campus_admin', 'course_admin']:
            user_campus_id = user.get('campus_id')
            batch_campus_ids = batch.get('campus_ids', [])
            
            # Handle both old and new batch structures
            if not batch_campus_ids and batch.get('campus_id'):
                batch_campus_ids = [batch.get('campus_id')]
            
            if not user_campus_id or ObjectId(user_campus_id) not in batch_campus_ids:
                return jsonify({'success': False, 'message': 'Access denied. You do not have permission to view this batch.'}), 403

        # Fetch campus and course names for the batch header
        campus_ids = batch.get('campus_ids', [])
        course_ids = batch.get('course_ids', [])
        
        campuses = list(mongo_db.campuses.find({'_id': {'$in': campus_ids}}))
        courses = list(mongo_db.courses.find({'_id': {'$in': course_ids}}))

        batch_info = {
            'id': str(batch['_id']),
            'name': batch.get('name'),
            'campus_name': ', '.join([c['name'] for c in campuses]),
            'course_name': ', '.join([c['name'] for c in courses]),
            'course_ids': [str(c['_id']) for c in courses],
        }

        # Fetch students with populated info using the new method
        students_with_details = mongo_db.get_students_by_batch(batch_id)

        return jsonify({
            'success': True,
            'data': students_with_details,
            'batch_info': batch_info
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batch students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/filtered', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_filtered_students():
    """Get filtered students with pagination"""
    try:
        current_app.logger.info("get_filtered_students endpoint called")
        print("get_filtered_students endpoint called")
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Get query parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        search = request.args.get('search', '')
        campus_id = request.args.get('campus_id', '')
        course_id = request.args.get('course_id', '')
        batch_id = request.args.get('batch_id', '')
        
        # Build query
        query = {'role': 'student'}
        
        if search:
            # We need to search in both users collection and students collection
            # First, find students with matching roll_number
            student_matches = list(mongo_db.students.find({
                'roll_number': {'$regex': search, '$options': 'i'}
            }))
            student_user_ids = [str(student['user_id']) for student in student_matches]
            
            # Build search query for users collection
            search_conditions = [
                {'name': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}}
            ]
            
            # Add user_id matches from students collection
            if student_user_ids:
                search_conditions.append({'_id': {'$in': [ObjectId(uid) for uid in student_user_ids]}})
            
            query['$or'] = search_conditions
        
        if campus_id:
            query['campus_id'] = ObjectId(campus_id)
        
        if course_id:
            query['course_id'] = ObjectId(course_id)
        
        if batch_id:
            query['batch_id'] = ObjectId(batch_id)
        
        # Super admin and sub_superadmin can see all students, others only their campus
        if user.get('role') not in ['superadmin', 'sub_superadmin']:
            user_campus_id = user.get('campus_id')
            if user_campus_id:
                query['campus_id'] = ObjectId(user_campus_id)
        
        # Get total count
        total = mongo_db.users.count_documents(query)
        
        # Get paginated results
        skip = (page - 1) * limit
        students = list(mongo_db.users.find(query).skip(skip).limit(limit))
        
        # Get additional student details
        student_details = []
        for student in students:
            # Get student profile
            student_profile = mongo_db.students.find_one({'user_id': student['_id']})
            
            # Get campus and course names
            campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')})
            course = mongo_db.courses.find_one({'_id': student.get('course_id')})
            batch = mongo_db.batches.find_one({'_id': student.get('batch_id')})
            
            student_details.append({
                '_id': str(student['_id']),
                'name': student.get('name', ''),
                'email': student.get('email', ''),
                'roll_number': student_profile.get('roll_number', '') if student_profile else '',
                'mobile_number': student_profile.get('mobile_number', '') if student_profile else '',
                'campus_name': campus.get('name', '') if campus else '',
                'course_name': course.get('name', '') if course else '',
                'batch_name': batch.get('name', '') if batch else '',
                'is_active': student.get('is_active', True),
                'created_at': student.get('created_at')
            })
        
        return jsonify({
            'success': True,
            'data': student_details,
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'has_more': (page * limit) < total
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching filtered students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/student/<student_id>', methods=['GET'])
@jwt_required()
def get_student_details(student_id):
    try:
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        user = mongo_db.users.find_one({'_id': student['user_id']})
        campus = mongo_db.campuses.find_one({'_id': student['campus_id']})
        course = mongo_db.courses.find_one({'_id': student['course_id']})
        batch = mongo_db.batches.find_one({'_id': student['batch_id']})

        student_details = {
            'id': str(student['_id']),
            'name': student['name'],
            'roll_number': student['roll_number'],
            'email': student['email'],
            'mobile_number': student['mobile_number'],
            'campus_name': campus['name'] if campus else 'N/A',
            'course_name': course['name'] if course else 'N/A',
            'batch_name': batch['name'] if batch else 'N/A',
            'username': user.get('username', 'N/A')
        }
        return jsonify({'success': True, 'data': student_details}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching student details: {e}")
        return jsonify({'success': False, 'message': 'An error occurred fetching student details.'}), 500

@batch_management_bp.route('/student/<student_id>', methods=['PUT'])
@jwt_required()
def update_student(student_id):
    try:
        data = request.json
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Uniqueness check for email and mobile_number (across all users except this one)
        email = data.get('email', student['email'])
        mobile_number = data.get('mobile_number', student['mobile_number'])
        user = mongo_db.users.find_one({'_id': student['user_id']})
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        existing_user = mongo_db.users.find_one({
            '_id': {'$ne': user['_id']},
            '$or': [
                {'email': email},
                {'mobile_number': mobile_number}
            ]
        })
        if existing_user:
            if existing_user['email'] == email:
                return jsonify({'success': False, 'message': 'Email already exists'}), 400
            return jsonify({'success': False, 'message': 'Mobile number already exists'}), 400

        # Update student collection
        student_update = {
            'name': data.get('name', student['name']),
            'roll_number': data.get('roll_number', student['roll_number']),
            'email': email,
            'mobile_number': mobile_number,
        }
        mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$set': student_update})

        # Update user collection
        user_update = {
            'name': data.get('name', student['name']),
            'email': email,
            'username': data.get('roll_number', student['roll_number']),
            'mobile_number': mobile_number
        }
        mongo_db.users.update_one({'_id': student['user_id']}, {'$set': user_update})

        return jsonify({'success': True, 'message': 'Student updated successfully'}), 200
    except Exception as e:
        current_app.logger.error(f"Error updating student: {e}")
        return jsonify({'success': False, 'message': 'An error occurred updating the student.'}), 500

@batch_management_bp.route('/student/<student_id>', methods=['DELETE'])
@jwt_required()
def delete_student(student_id):
    try:
        student = mongo_db.students.find_one_and_delete({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Also delete the associated user account
        mongo_db.users.delete_one({'_id': student['user_id']})

        return jsonify({'success': True, 'message': 'Student deleted successfully'}), 200
    except Exception as e:
        current_app.logger.error(f"Error deleting student: {e}")
        return jsonify({'success': False, 'message': 'An error occurred deleting the student.'}), 500

@batch_management_bp.route('/student/<student_id>/authorize-level', methods=['POST'])
@jwt_required()
def authorize_student_level(student_id):
    try:
        data = request.json
        level = data.get('level')
        if not level:
            return jsonify({'success': False, 'message': 'Level is required'}), 400
        # Ensure authorized_levels exists
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404
        if 'authorized_levels' not in student:
            mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$set': {'authorized_levels': []}})
        # Add the level to authorized_levels
        result = mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$addToSet': {'authorized_levels': level}})
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if result.modified_count == 0:
            return jsonify({'success': False, 'message': f"Level '{level}' was already authorized for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
        return jsonify({'success': True, 'message': f"Level '{level}' authorized for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error authorizing level: {e}")
        return jsonify({'success': False, 'message': 'An error occurred authorizing the level.'}), 500

@batch_management_bp.route('/student/<student_id>/lock-level', methods=['POST'])
@jwt_required()
def lock_student_level(student_id):
    try:
        data = request.json
        level = data.get('level')
        if not level:
            return jsonify({'success': False, 'message': 'Level is required'}), 400

        # Find student by user_id or _id
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404

        # Remove level from authorized_levels
        mongo_db.students.update_one({'_id': student['_id']}, {'$pull': {'authorized_levels': level}})
        student = mongo_db.students.find_one({'_id': student['_id']})

        # Emit real-time event to the student
        socketio.emit('level_access_changed', {'student_id': str(student['_id']), 'level': level, 'action': 'locked'}, room=str(student['_id']))

        return jsonify({'success': True, 'message': f"Level '{level}' locked for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error locking level: {e}")
        return jsonify({'success': False, 'message': 'An error occurred locking the level.'}), 500

@batch_management_bp.route('/student/<student_id>/authorize-module', methods=['POST'])
@batch_management_bp.route('/student/<student_id>/authorize-module/', methods=['POST'])
@jwt_required()
def authorize_student_module(student_id):
    try:
        # Support multiple input forms: JSON body, form data, and alternate keys
        data = request.get_json(silent=True) or {}
        # fallback to form data if request.json is None or empty
        if not data and request.form:
            data = request.form.to_dict()

        # Accept module from multiple possible keys for robustness
        module = data.get('module') or data.get('module_id') or data.get('moduleId')
        reason = data.get('reason', 'Admin authorization')

        # Debug logging: capture incoming payload and module for easier troubleshooting
        current_app.logger.info(f"authorize_student_module called - student_id={student_id}, module={module}, payload={data}")

        if not module:
            # Log raw request body and headers for debugging
            try:
                raw = request.get_data(as_text=True)
            except Exception:
                raw = '<unable to read body>'
            current_app.logger.warning(f"authorize_student_module missing 'module' - raw_body={raw} headers={dict(request.headers)}")
            return jsonify({'success': False, 'message': 'Module is required'}), 400

        # Get current admin user
        current_user_id = get_jwt_identity()
        current_app.logger.info(f"authorize_student_module: current_user_id={current_user_id}")

        # Find student - try multiple methods
        student = None
        try:
            # Method 1: Try as ObjectId (student._id)
            try:
                student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
            except Exception:
                student = None

            # Method 2: Try as user_id
            if not student:
                try:
                    student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
                except Exception:
                    student = None

            # Method 3: Try as roll_number
            if not student:
                student = mongo_db.students.find_one({'roll_number': student_id})

            # Method 4: Try as username in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'username': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})

            # Method 5: Try as email in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'email': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})

            if not student:
                current_app.logger.warning(f"authorize_student_module: student not found for identifier {student_id}")
                return jsonify({'success': False, 'message': f'Student not found with identifier: {student_id}'}), 404

        except Exception as e:
            current_app.logger.exception(f"Error searching for student: {e}")
            return jsonify({'success': False, 'message': f'Error searching for student: {str(e)}'}), 400

        # Compute module levels (quick sanity) and log
        from config.constants import LEVELS
        module_levels_check = [level_id for level_id, level in LEVELS.items() if isinstance(level, dict) and (level.get('module_id') == module or level.get('module') == module)]
        current_app.logger.info(f"authorize_student_module: module_levels_check for module={module} -> {module_levels_check}")

        # Use progress manager for enhanced authorization
        from utils.student_progress_manager import StudentProgressManager
        progress_manager = StudentProgressManager(mongo_db)

        try:
            # Ensure we pass a string id to the progress manager (it expects ids it can wrap with ObjectId)
            success, message = progress_manager.admin_authorize_module(
                student_id=str(student['_id']),
                module_id=module,
                admin_user_id=current_user_id,
                reason=reason
            )
            current_app.logger.info(f"authorize_student_module: progress_manager returned success={success}, message={message}")
        except Exception as pm_err:
            current_app.logger.exception(f"authorize_student_module: progress_manager raised exception: {pm_err}")
            return jsonify({'success': False, 'message': f'Error authorizing module: {str(pm_err)}'}), 500

        if success:
            # Emit real-time event to the student using socketio instance
            try:
                socketio.emit('module_access_changed', {
                    'student_id': str(student['_id']),
                    'module': module,
                    'action': 'unlocked',
                    'authorized_by': 'admin'
                }, room=str(student['_id']))
            except Exception as socket_error:
                current_app.logger.warning(f"Could not emit socket event: {socket_error}")

            # Get updated student data
            updated_student = mongo_db.students.find_one({'_id': student['_id']})

            return jsonify({
                'success': True,
                'message': message,
                'authorized_levels': updated_student.get('authorized_levels', [])
            }), 200
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        current_app.logger.exception(f"Error authorizing module: {e}")
        return jsonify({'success': False, 'message': 'An error occurred authorizing the module.'}), 500

@batch_management_bp.route('/student/<student_id>/lock-module', methods=['POST'])
@jwt_required()
def lock_student_module(student_id):
    try:
        data = request.json
        module = data.get('module')
        if not module:
            return jsonify({'success': False, 'message': 'Module is required'}), 400

        from config.constants import LEVELS
        module_levels = [level_id for level_id, level in LEVELS.items() if level.get('module_id') == module or level.get('module') == module]
        if not module_levels:
            return jsonify({'success': False, 'message': 'No levels found for this module.'}), 404

        # Remove all levels from authorized_levels
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404
        # Pull object entries first, then string entries to support legacy and new shapes
        try:
            mongo_db.students.update_one(
                {'_id': student['_id']},
                {'$pull': {'authorized_levels': {'level_id': {'$in': module_levels}}}}
            )
        except Exception as e:
            current_app.logger.exception(f"Failed to pull object-authorized_levels: {e}")

        try:
            mongo_db.students.update_one(
                {'_id': student['_id']},
                {'$pull': {'authorized_levels': {'$in': module_levels}}}
            )
        except Exception as e:
            current_app.logger.exception(f"Failed to pull string-authorized_levels: {e}")

        student = mongo_db.students.find_one({'_id': student['_id']})

        current_app.logger.info(f"lock_student_module: removed levels for module={module}, module_levels={module_levels}, remaining_authorized_levels_count={len(student.get('authorized_levels', []))}")

        # Emit real-time event to the student
        socketio.emit('module_access_changed', {'student_id': str(student['_id']), 'module': module, 'action': 'locked'}, room=str(student['_id']))

        return jsonify({'success': True, 'message': f"Module '{module}' locked for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error locking module: {e}")
        return jsonify({'success': False, 'message': 'An error occurred locking the module.'}), 500

@batch_management_bp.route('/create-with-students', methods=['POST'])
@jwt_required()
def create_batch_with_students():
    try:
        data = request.get_json()
        name = data.get('name')
        campus_ids = [ObjectId(cid) for cid in data.get('campus_ids', [])]
        course_ids = [ObjectId(cid) for cid in data.get('course_ids', [])]
        students_data = data.get('students', [])

        if not name or not campus_ids or not students_data:
            return jsonify({'success': False, 'message': 'Batch name, campus, and student data are required.'}), 400
        
        if mongo_db.batches.find_one({'name': name}):
            return jsonify({'success': False, 'message': 'A batch with this name already exists.'}), 409

        # 1. Create the batch
        batch_doc = {
            'name': name,
            'campus_ids': campus_ids,
            'course_ids': course_ids,
            'created_at': datetime.now(pytz.utc)
        }
        batch_id = mongo_db.batches.insert_one(batch_doc).inserted_id

        # 2. Create students and users
        created_students_details = []
        errors = []
        
        for student in students_data:
            try:
                # Find campus by name from the uploaded file
                campus = mongo_db.campuses.find_one({'name': student['campus_name']})
                if not campus:
                    errors.append(f"Campus '{student['campus_name']}' not found for student '{student.get('student_name', 'N/A')}'.")
                    continue
                
                # Find course by name and campus_id
                course = mongo_db.courses.find_one({
                    'name': student['course_name'],
                    'campus_id': campus['_id']
                })
                if not course:
                    errors.append(f"Course '{student['course_name']}' not found in campus '{student['campus_name']}' for student '{student.get('student_name', 'N/A')}'.")
                    continue

                # Check for roll number duplicates (still enforced)
                existing_user = mongo_db.users.find_one({'username': student['roll_number']})
                if existing_user:
                    errors.append(f"Student with roll number '{student['roll_number']}' already exists.")
                    continue
                
                # Check for email duplicates (now just a warning)
                if student.get('email'):
                    existing_email_user = mongo_db.users.find_one({'email': student['email']})
                    if existing_email_user:
                        current_app.logger.warning(f"⚠️ Email {student['email']} already exists for another user, but allowing duplicate")

                # Check for duplicate mobile number if provided
                if student.get('mobile_number') and mongo_db.users.find_one({'mobile_number': student['mobile_number']}):
                    errors.append(f"Student with mobile number '{student['mobile_number']}' already exists.")
                    continue

                username = student['roll_number']
                password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

                user_doc = {
                    'username': username,
                    'email': student['email'],
                    'password_hash': password_hash,
                    'role': ROLES['STUDENT'],
                    'name': student['student_name'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_id,
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                user_id = mongo_db.users.insert_one(user_doc).inserted_id

                student_doc = {
                    'user_id': user_id,
                    'name': student['student_name'],
                    'roll_number': student['roll_number'],
                    'email': student['email'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_id,
                    'created_at': datetime.now(pytz.utc)
                }
                mongo_db.students.insert_one(student_doc)

                created_students_details.append({
                    "student_name": student['student_name'],
                    "email": student['email'],
                    "username": username,
                    "password": password
                })

            except Exception as student_error:
                errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")

        # 3. Send emails with progress updates
        current_user_id = get_jwt_identity()
        total_emails = len(created_students_details)
        
        # Send initial progress update
        socketio.emit('upload_progress', {
            'user_id': current_user_id,
            'status': 'sending_emails',
            'total': total_emails,
            'processed': 0,
            'percentage': 0,
            'message': 'Sending welcome emails to students...'
        }, room=str(current_user_id))
        
        for index, student_details in enumerate(created_students_details):
             email_sent = False
             email_error = None
             try:
                html_content = render_template('student_credentials.html', params={
                    'name': student_details['student_name'],
                    'username': student_details['username'],
                    'email': student_details['email'],
                    'password': student_details['password'],
                    'login_url': "https://crt.pydahsoft.in/login"
                })
                send_email(to_email=student_details['email'], to_name=student_details['student_name'], subject="Welcome to VERSANT - Your Student Credentials", html_content=html_content)
                email_sent = True
                
             except Exception as email_error:
                email_error = str(email_error)
                current_app.logger.error(f"Failed to send welcome email to {student_details['email']}: {email_error}")
                # Don't add to errors array - just log it
                
             # Send progress update after email attempt (regardless of success/failure)
             percentage = int(((index + 1) / total_emails) * 100)
             if email_sent:
                 socketio.emit('upload_progress', {
                     'user_id': current_user_id,
                     'status': 'sending_emails',
                     'total': total_emails,
                     'processed': index + 1,
                     'percentage': percentage,
                     'message': f'Email sent to {student_details["student_name"]} ({student_details["email"]})',
                     'current_student': {
                         'name': student_details['student_name'],
                         'email': student_details['email'],
                         'username': student_details['username']
                     }
                 }, room=str(current_user_id))
             else:
                 socketio.emit('upload_progress', {
                     'user_id': current_user_id,
                     'status': 'sending_emails',
                     'total': total_emails,
                     'processed': index + 1,
                     'percentage': percentage,
                     'message': f'Email sending failed for {student_details["student_name"]} ({student_details["email"]}) - Student created successfully',
                     'current_student': {
                         'name': student_details['student_name'],
                         'email': student_details['email'],
                         'username': student_details['username']
                     },
                     'email_warning': True,
                     'email_error': email_error
                 }, room=str(current_user_id))
        
        if errors:
            return jsonify({
                'success': False, 
                'message': f"Batch '{name}' created, but some students could not be added.", 
                'data': {'batch_id': str(batch_id), 'created_students': created_students_details}, 
                'errors': errors
            }), 207
        
        return jsonify({
            'success': True, 
            'message': f"Batch '{name}' created successfully with {len(created_students_details)} students.", 
            'data': {'batch_id': str(batch_id), 'created_students': created_students_details}
        }), 201

    except Exception as e:
        current_app.logger.error(f"Error creating batch with students: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>/add-students', methods=['POST'])
@jwt_required()
def add_students_to_batch(batch_id):
    try:
        # Support both file upload (preferred) and JSON (legacy)
        if 'student_file' in request.files:
            file = request.files['student_file']
            batch_obj_id = ObjectId(batch_id)
            batch = mongo_db.batches.find_one({'_id': batch_obj_id})
            if not batch:
                return jsonify({'success': False, 'message': 'Batch not found.'}), 404
            campus_ids = batch.get('campus_ids', [])
            course_ids = batch.get('course_ids', [])
            if not campus_ids or not course_ids:
                return jsonify({'success': False, 'message': 'Batch is missing campus or course info.'}), 400
            campus_id = campus_ids[0]  # Assume single campus per batch for now
            # Parse and validate file
            rows = _parse_student_file(file)
            if not rows:
                return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400
            # Get campus and course info
            campus = mongo_db.campuses.find_one({'_id': campus_id})
            valid_course_names = set(c['name'] for c in mongo_db.courses.find({'_id': {'$in': course_ids}}))
            # Fetch existing data for validation
            existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
            # Note: Email duplicates are now allowed, so we don't need to check existing emails
            existing_emails = set()  # Empty set since we allow email duplicates
            existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))
            preview_data = []
            for row in rows:
                student_data = {
                    'campus_name': str(row.get('Campus Name', '')).strip(),
                    'course_name': str(row.get('Course Name', '')).strip(),
                    'student_name': str(row.get('Student Name', '')).strip(),
                    'roll_number': str(row.get('Roll Number', '')).strip(),
                    'email': str(row.get('Email', '')).strip().lower(),
                    'mobile_number': str(row.get('Mobile Number', '')).strip(),
                }
                errors = []
                if not all([student_data['campus_name'], student_data['course_name'], student_data['student_name'], student_data['roll_number'], student_data['email']]):
                    errors.append('Missing required fields.')
                if student_data['roll_number'] in existing_roll_numbers:
                    errors.append('Roll number already exists.')
                if student_data['email'] in existing_emails:
                    errors.append('Email already exists.')
                if student_data['mobile_number'] and student_data['mobile_number'] in existing_mobile_numbers:
                    errors.append('Mobile number already exists.')
                if student_data['campus_name'] != campus['name']:
                    errors.append(f"Campus '{student_data['campus_name']}' doesn't match batch campus '{campus['name']}'.")
                if student_data['course_name'] not in valid_course_names:
                    errors.append(f"Course '{student_data['course_name']}' not found in this batch.")
                student_data['errors'] = errors
                preview_data.append(student_data)
            # Only add students with no errors
            created_students_details = []
            errors = []
            uploaded_emails = []  # Track emails for verification
            
            for student in preview_data:
                if student['errors']:
                    errors.append(f"{student['student_name']}: {', '.join(student['errors'])}")
                    continue
                try:
                    course = mongo_db.courses.find_one({'name': student['course_name'], '_id': {'$in': course_ids}})
                    username = student['roll_number']
                    password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                    
                    user_doc = {
                        'username': username,
                        'email': student['email'],
                        'password_hash': password_hash,
                        'role': ROLES['STUDENT'],
                        'name': student['student_name'],
                        'mobile_number': student.get('mobile_number', ''),
                        'campus_id': campus_id,
                        'course_id': course['_id'],
                        'batch_id': batch_obj_id,
                        'is_active': True,
                        'created_at': datetime.now(pytz.utc),
                        'mfa_enabled': False
                    }
                    
                    # Insert user first
                    user_result = mongo_db.users.insert_one(user_doc)
                    if not user_result.inserted_id:
                        errors.append(f"{student['student_name']}: Failed to create user account.")
                        continue
                    
                    user_id = user_result.inserted_id
                    
                    student_doc = {
                        'user_id': user_id,
                        'name': student['student_name'],
                        'roll_number': student['roll_number'],
                        'email': student['email'],
                        'mobile_number': student.get('mobile_number', ''),
                        'campus_id': campus_id,
                        'course_id': course['_id'],
                        'batch_id': batch_obj_id,
                        'created_at': datetime.now(pytz.utc)
                    }
                    
                    # Insert student profile
                    student_result = mongo_db.students.insert_one(student_doc)
                    if not student_result.inserted_id:
                        # Rollback user creation if student creation fails
                        mongo_db.users.delete_one({'_id': user_id})
                        errors.append(f"{student['student_name']}: Failed to create student profile.")
                        continue
                    
                    created_students_details.append({
                        "student_name": student['student_name'],
                        "email": student['email'],
                        "username": username,
                        "password": password
                    })
                    uploaded_emails.append(student['email'])
                    
                    # Send welcome email
                    try:
                        html_content = render_template('student_credentials.html', params={
                            'name': student['student_name'],
                            'username': username,
                            'email': student['email'],
                            'password': password,
                            'login_url': "https://crt.pydahsoft.in/login"
                        })
                        send_email(to_email=student['email'], to_name=student['student_name'], subject="Welcome to Study Edge - Your Student Credentials", html_content=html_content)
                    except Exception as e:
                        errors.append(f"Failed to send email to {student['email']}: {e}")
                        
                except Exception as student_error:
                    errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")
            
            # Verify upload success
            verification_results = []
            if uploaded_emails:
                students_verified = list(mongo_db.students.find({'email': {'$in': uploaded_emails}, 'batch_id': batch_obj_id}))
                users_verified = list(mongo_db.users.find({'email': {'$in': uploaded_emails}, 'batch_id': batch_obj_id}))
                
                for email in uploaded_emails:
                    student_exists = any(s['email'] == email for s in students_verified)
                    user_exists = any(u['email'] == email for u in users_verified)
                    verification_results.append({
                        'email': email,
                        'student_profile_exists': student_exists,
                        'user_account_exists': user_exists,
                        'fully_uploaded': student_exists and user_exists
                    })
            
            if errors:
                return jsonify({
                    'success': bool(created_students_details), 
                    'message': f"Process completed with {len(errors)} errors.", 
                    'data': {
                        'created_students': created_students_details,
                        'verification_results': verification_results
                    }, 
                    'errors': errors
                }), 207
            
            return jsonify({
                'success': True, 
                'message': f"Successfully added {len(created_students_details)} students to the batch.", 
                'data': {
                    'created_students': created_students_details,
                    'verification_results': verification_results
                }
            }), 201
        # Fallback: legacy JSON method
        data = request.get_json()
        students_data = data.get('students', []) if data else []
        if not students_data:
            return jsonify({'success': False, 'message': 'Student data are required.'}), 400
        batch_obj_id = ObjectId(batch_id)
        batch = mongo_db.batches.find_one({'_id': batch_obj_id})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        valid_campus_ids = batch.get('campus_ids', [])
        valid_course_ids = batch.get('course_ids', [])
        created_students_details = []
        errors = []
        for student in students_data:
            try:
                campus = mongo_db.campuses.find_one({'name': student['campus_name']})
                if not campus or campus['_id'] not in valid_campus_ids:
                    errors.append(f"Campus '{student['campus_name']}' is not valid for this batch for student '{student.get('student_name', 'N/A')}'.")
                    continue
                course = mongo_db.courses.find_one({'name': student['course_name'], 'campus_id': campus['_id']})
                if not course or course['_id'] not in valid_course_ids:
                    errors.append(f"Course '{student['course_name']}' is not valid for this batch for student '{student.get('student_name', 'N/A')}'.")
                    continue
                # Check for roll number duplicates (still enforced)
                existing_user = mongo_db.users.find_one({'username': student['roll_number']})
                if existing_user:
                    errors.append(f"Student with roll number '{student['roll_number']}' already exists.")
                    continue
                
                # Check for email duplicates (now just a warning)
                if student.get('email'):
                    existing_email_user = mongo_db.users.find_one({'email': student['email']})
                    if existing_email_user:
                        current_app.logger.warning(f"⚠️ Email {student['email']} already exists for another user, but allowing duplicate")
                # Check for duplicate mobile number if provided
                if student.get('mobile_number') and mongo_db.users.find_one({'mobile_number': student['mobile_number']}):
                    errors.append(f"Student with mobile number '{student['mobile_number']}' already exists.")
                    continue
                username = student['roll_number']
                password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                user_doc = {
                    'username': username,
                    'email': student['email'],
                    'password_hash': password_hash,
                    'role': ROLES['STUDENT'],
                    'name': student['student_name'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_obj_id,
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                user_id = mongo_db.users.insert_one(user_doc).inserted_id
                student_doc = {
                    'user_id': user_id,
                    'name': student['student_name'],
                    'roll_number': student['roll_number'],
                    'email': student['email'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_obj_id,
                    'created_at': datetime.now(pytz.utc)
                }
                mongo_db.students.insert_one(student_doc)
                created_students_details.append({
                    "student_name": student['student_name'],
                    "email": student['email'],
                    "username": username,
                    "password": password
                })
                # Send welcome email
                try:
                    html_content = render_template('student_credentials.html', params={
                        'name': student['student_name'],
                        'username': username,
                        'email': student['email'],
                        'password': password,
                        'login_url': "https://crt.pydahsoft.in/login"
                    })
                    send_email(to_email=student['email'], to_name=student['student_name'], subject="Welcome to VERSANT - Your Student Credentials", html_content=html_content)
                except Exception as e:
                    errors.append(f"Failed to send email to {student['email']}: {e}")
            except Exception as student_error:
                errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")
        if errors:
            return jsonify({'success': bool(created_students_details), 'message': f"Process completed with {len(errors)} errors.", 'data': {'created_students': created_students_details}, 'errors': errors}), 207
        return jsonify({'success': True, 'message': f"Successfully added {len(created_students_details)} students to the batch.", 'data': {'created_students': created_students_details}}), 201
    except Exception as e:
        current_app.logger.error(f"Error adding students to batch: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/add-student', methods=['POST'])
@jwt_required()
def add_single_student():
    try:
        data = request.get_json()
        batch_id = data.get('batch_id')
        course_id = data.get('course_id')
        name = data.get('name')
        roll_number = data.get('roll_number')
        email = data.get('email')
        mobile_number = data.get('mobile_number')

        if not all([batch_id, course_id, name, roll_number, email]):
            return jsonify({'success': False, 'message': 'All fields are required.'}), 400

        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        campus_ids = batch.get('campus_ids', [])
        if not campus_ids:
            return jsonify({'success': False, 'message': 'Batch is missing campus info.'}), 400
        campus_id = campus_ids[0]  # Assume single campus per batch for now

        # Validate course_id
        valid_course_ids = set(str(cid) for cid in batch.get('course_ids', []))
        if course_id not in valid_course_ids:
            return jsonify({'success': False, 'message': 'Course is not valid for this batch.'}), 400

        # Check for existing user
        if mongo_db.users.find_one({'username': roll_number}):
            return jsonify({'success': False, 'message': 'Roll number already exists.'}), 400
        if mongo_db.users.find_one({'email': email}):
            return jsonify({'success': False, 'message': 'Email already exists.'}), 400
        if mobile_number and mongo_db.users.find_one({'mobile_number': mobile_number}):
            return jsonify({'success': False, 'message': 'Mobile number already exists.'}), 400

        # Create user and student
        password = f"{name.split()[0][:4].lower()}{roll_number[-4:]}"
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        user_doc = {
            'username': roll_number,
            'email': email,
            'password_hash': password_hash,
            'role': 'student',
            'name': name,
            'mobile_number': mobile_number,
            'campus_id': campus_id,
            'course_id': ObjectId(course_id),
            'batch_id': ObjectId(batch_id),
            'is_active': True,
            'created_at': datetime.now(pytz.utc),
            'mfa_enabled': False
        }
        user_id = mongo_db.users.insert_one(user_doc).inserted_id
        student_doc = {
            'user_id': user_id,
            'name': name,
            'roll_number': roll_number,
            'email': email,
            'mobile_number': mobile_number,
            'campus_id': campus_id,
            'course_id': ObjectId(course_id),
            'batch_id': ObjectId(batch_id),
            'created_at': datetime.now(pytz.utc)
        }
        mongo_db.students.insert_one(student_doc)
        # Send welcome email
        try:
            html_content = render_template(
                'student_credentials.html',
                params={
                    'name': name,
                    'username': roll_number,
                    'email': email,
                    'password': password,
                    'login_url': "https://crt.pydahsoft.in/login"
                }
            )
            send_email(
                to_email=email,
                to_name=name,
                subject="Welcome to VERSANT - Your Student Credentials",
                html_content=html_content
            )
        except Exception as e:
            return jsonify({'success': True, 'message': 'Student added, but failed to send email.', 'created_students': [{
                'student_name': name,
                'username': roll_number,
                'password': password,
                'email': email
            }], 'email_error': str(e)}), 200
        return jsonify({'success': True, 'message': 'Student added successfully!', 'created_students': [{
            'student_name': name,
            'username': roll_number,
            'password': password,
            'email': email
        }]}), 200
    except Exception as e:
        current_app.logger.error(f"Error adding single student: {str(e)}")
        return jsonify({'success': False, 'message': f'An unexpected server error occurred: {str(e)}'}), 500

@batch_management_bp.route('/student/<student_id>/access-status', methods=['GET'])
@jwt_required()
def get_student_access_status(student_id):
    try:
        from config.constants import MODULES, LEVELS
        
        # Find student - try multiple methods
        student = None
        try:
            # Method 1: Try as ObjectId (student._id)
            try:
                student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
            except:
                pass
            
            # Method 2: Try as user_id
            if not student:
                try:
                    student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
                except:
                    pass
            
            # Method 3: Try as roll_number
            if not student:
                student = mongo_db.students.find_one({'roll_number': student_id})
            
            # Method 4: Try as username in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'username': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})
            
            # Method 5: Try as email in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'email': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})
                    
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error searching for student: {str(e)}'}), 400
        
        # Default logic if student not found or no authorized_levels
        default_grammar_unlocked = ['GRAMMAR_NOUN']
        modules_status = []
        if not student or not student.get('authorized_levels'):
            for module_id, module_name in MODULES.items():
                levels = [
                    {
                        'level_id': level_id,
                        'level_name': level['name'] if isinstance(level, dict) else level,
                        'unlocked': (
                            (module_id == 'GRAMMAR' and level_id == 'GRAMMAR_NOUN') or
                            (module_id == 'VOCABULARY')
                        )
                    }
                    for level_id, level in LEVELS.items()
                    if (level.get('module_id') if isinstance(level, dict) else None) == module_id
                ]
                modules_status.append({
                    'module_id': module_id,
                    'module_name': module_name,
                    'unlocked': (
                        module_id == 'GRAMMAR' or module_id == 'VOCABULARY'
                    ),
                    'levels': levels
                })
            return jsonify({'success': True, 'data': modules_status}), 200
        # If student has authorized_levels, use them
        authorized_levels_raw = student.get('authorized_levels', [])
        
        # Handle both old format (strings) and new format (objects)
        authorized_levels = set()
        if authorized_levels_raw:
            if isinstance(authorized_levels_raw[0], str):
                # Old format - just strings
                authorized_levels = set(authorized_levels_raw)
            else:
                # New format - objects with metadata
                authorized_levels = set([level['level_id'] for level in authorized_levels_raw])
        
        for module_id, module_name in MODULES.items():
            levels = [
                {
                    'level_id': level_id,
                    'level_name': level['name'] if isinstance(level, dict) else level,
                    'unlocked': level_id in authorized_levels
                }
                for level_id, level in LEVELS.items()
                if (level.get('module_id') if isinstance(level, dict) else None) == module_id
            ]
            modules_status.append({
                'module_id': module_id,
                'module_name': module_name,
                'unlocked': all(l['unlocked'] for l in levels) if levels else False,
                'levels': levels
            })
        return jsonify({'success': True, 'data': modules_status}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching access status: {e}")
        return jsonify({'success': False, 'message': 'An error occurred fetching access status.'}), 500

@batch_management_bp.route('/student/<student_id>/detailed-insights', methods=['GET'])
@jwt_required()
def get_student_detailed_insights(student_id):
    """Get comprehensive student insights for admin dashboard"""
    try:
        # Get current admin user
        current_user_id = get_jwt_identity()
        user = mongo_db.users.find_one({'_id': ObjectId(current_user_id)})
        
        # Check admin permissions
        allowed_roles = ['super_admin', 'superadmin', 'campus_admin', 'course_admin', 'sub_superadmin']
        if not user:
            return jsonify({'success': False, 'message': 'User not found.'}), 403
        
        user_role = user.get('role')
        if user_role not in allowed_roles:
            return jsonify({'success': False, 'message': f'Access denied. Required roles: {allowed_roles}, but user has: {user_role}'}), 403
        
        # Find student - try multiple methods
        student = None
        try:
            # Method 1: Try as ObjectId (student._id)
            try:
                student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
            except:
                pass
            
            # Method 2: Try as user_id
            if not student:
                try:
                    student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
                except:
                    pass
            
            # Method 3: Try as roll_number
            if not student:
                student = mongo_db.students.find_one({'roll_number': student_id})
            
            # Method 4: Try as username in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'username': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})
            
            # Method 5: Try as email in users collection, then find student
            if not student:
                user = mongo_db.users.find_one({'email': student_id})
                if user:
                    student = mongo_db.students.find_one({'user_id': user['_id']})
            
            if not student:
                return jsonify({'success': False, 'message': f'Student not found with identifier: {student_id}'}), 404
                
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error searching for student: {str(e)}'}), 400
        
        # Use progress manager for detailed insights
        try:
            from utils.student_progress_manager import StudentProgressManager
            progress_manager = StudentProgressManager(mongo_db)
            
            current_app.logger.info(f"Generating insights for student: {student['_id']}")
            insights = progress_manager.get_student_detailed_insights(student['_id'])
            
            if insights:
                current_app.logger.info(f"Insights generated successfully for student: {student.get('name', 'Unknown')}")
                
                # Convert ObjectIds to strings for JSON serialization
                def convert_objectids(obj):
                    if isinstance(obj, ObjectId):
                        return str(obj)
                    elif isinstance(obj, dict):
                        return {key: convert_objectids(value) for key, value in obj.items()}
                    elif isinstance(obj, list):
                        return [convert_objectids(item) for item in obj]
                    elif hasattr(obj, 'isoformat'):  # datetime objects
                        return obj.isoformat()
                    else:
                        return obj
                
                serializable_insights = convert_objectids(insights)
                return jsonify({'success': True, 'data': serializable_insights}), 200
            else:
                current_app.logger.warning(f"Failed to generate insights for student: {student['_id']}")
                return jsonify({'success': False, 'message': 'Failed to generate insights.'}), 500
        except Exception as e:
            current_app.logger.error(f"Error generating insights: {e}", exc_info=True)
            return jsonify({'success': False, 'message': f'Error generating insights: {str(e)}'}), 500
            
    except Exception as e:
        current_app.logger.error(f"Error getting student insights: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'An error occurred fetching student insights: {str(e)}'}), 500

@batch_management_bp.route('/system/progress-monitoring', methods=['GET'])
@jwt_required()
def get_progress_system_monitoring():
    """Get system monitoring data for progress management"""
    try:
        # Get current admin user
        current_user_id = get_jwt_identity()
        user = mongo_db.users.find_one({'_id': ObjectId(current_user_id)})
        
        # Check admin permissions (only super_admin and campus_admin)
        allowed_roles = ['super_admin', 'campus_admin','sub_superadmin']
        if not user or user.get('role') not in allowed_roles:
            return jsonify({'success': False, 'message': 'Access denied. Super admin privileges required.'}), 403
        
        # Get monitoring data
        from utils.student_progress_manager import StudentProgressManager
        progress_manager = StudentProgressManager(mongo_db)
        
        # Get system health metrics
        health_metrics = progress_manager.monitoring.get_system_health_metrics()
        
        # Get progress analytics
        analytics = progress_manager.monitoring.get_student_progress_analytics(days=30)
        
        # Get data integrity status
        integrity_check = progress_manager.monitoring.validate_student_progress_integrity()
        
        monitoring_data = {
            'health_metrics': health_metrics,
            'analytics': analytics,
            'integrity_check': integrity_check,
            'timestamp': datetime.utcnow().isoformat()
        }
        
        return jsonify({'success': True, 'data': monitoring_data}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error getting progress monitoring data: {e}")
        return jsonify({'success': False, 'message': 'An error occurred fetching monitoring data.'}), 500

# --- BATCH-COURSE INSTANCE MANAGEMENT ---

@batch_management_bp.route('/instances', methods=['GET'])
@jwt_required()
def get_batch_course_instances():
    """Get all batch-course instances with details"""
    try:
        pipeline = [
            {
                '$lookup': {
                    'from': 'batches',
                    'localField': 'batch_id',
                    'foreignField': '_id',
                    'as': 'batch'
                }
            },
            {
                '$lookup': {
                    'from': 'courses',
                    'localField': 'course_id',
                    'foreignField': '_id',
                    'as': 'course'
                }
            },
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'batch.campus_ids',
                    'foreignField': '_id',
                    'as': 'campuses'
                }
            },
            {
                '$unwind': '$batch'
            },
            {
                '$unwind': '$course'
            },
            {
                '$project': {
                    '_id': 1,
                    'batch_name': '$batch.name',
                    'course_name': '$course.name',
                    'campus_names': '$campuses.name',
                    'student_count': {'$size': {'$ifNull': ['$students', []]}},
                    'created_at': 1
                }
            },
            {'$sort': {'created_at': -1}}
        ]
        
        instances = list(mongo_db.db.batch_course_instances.aggregate(pipeline))
        
        # Convert ObjectIds to strings
        for instance in instances:
            instance['_id'] = str(instance['_id'])
            if 'created_at' in instance:
                instance['created_at'] = safe_isoformat(instance['created_at'])
        
        return jsonify({'success': True, 'data': instances}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batch-course instances: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>', methods=['GET'])
@jwt_required()
def get_batch_course_instance_details(instance_id):
    """Get detailed information about a specific batch-course instance"""
    try:
        instance = mongo_db.db.batch_course_instances.find_one({'_id': ObjectId(instance_id)})
        if not instance:
            return jsonify({'success': False, 'message': 'Instance not found'}), 404
        
        # Get batch details
        batch = mongo_db.batches.find_one({'_id': instance['batch_id']})
        course = mongo_db.courses.find_one({'_id': instance['course_id']})
        
        # Get students in this instance
        students = list(mongo_db.students.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        # Get test results for this instance
        test_results = list(mongo_db.student_test_attempts.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        instance_data = {
            'id': str(instance['_id']),
            'batch': {
                'id': str(batch['_id']),
                'name': batch['name']
            } if batch else None,
            'course': {
                'id': str(course['_id']),
                'name': course['name']
            } if course else None,
            'student_count': len(students),
            'test_results_count': len(test_results),
            'created_at': safe_isoformat(instance.get('created_at')) if instance.get('created_at') else None
        }
        
        return jsonify({'success': True, 'data': instance_data}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching instance details: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>/students', methods=['GET'])
@jwt_required()
def get_instance_students(instance_id):
    """Get all students in a specific batch-course instance"""
    try:
        students = list(mongo_db.students.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        student_list = []
        for student in students:
            # Get user details
            user = mongo_db.users.find_one({'_id': student['user_id']})
            student_list.append({
                'id': str(student['_id']),
                'name': student['name'],
                'roll_number': student['roll_number'],
                'email': student['email'],
                'username': user['username'] if user else '',
                'is_active': user['is_active'] if user else True,
                'created_at': safe_isoformat(student['created_at']) if student.get('created_at') else None
            })
        
        return jsonify({'success': True, 'data': student_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching instance students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>/upload-students', methods=['POST'])
@jwt_required()
def upload_students_to_instance(instance_id):
    """Upload students to a specific batch-course instance"""
    try:
        # Verify instance exists
        instance = mongo_db.db.batch_course_instances.find_one({'_id': ObjectId(instance_id)})
        if not instance:
            return jsonify({'success': False, 'message': 'Instance not found'}), 404
        
        # Get batch and course details
        batch = mongo_db.batches.find_one({'_id': instance['batch_id']})
        course = mongo_db.courses.find_one({'_id': instance['course_id']})
        
        if not batch or not course:
            return jsonify({'success': False, 'message': 'Batch or course not found'}), 404
        
        # Handle file upload
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Parse student file
        rows = _parse_student_file(file)
        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid'}), 400
        
        # Validate columns
        columns = list(rows[0].keys()) if rows else []
        required_fields = ['Student Name', 'Roll Number', 'Email', 'Mobile Number']
        missing_fields = [field for field in required_fields if field not in columns]
        if missing_fields:
            return jsonify({'success': False, 'message': f"Invalid file structure. Missing columns: {', '.join(missing_fields)}"}), 400
        
        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        # Note: Email duplicates are now allowed, so we don't need to check existing emails
        existing_emails = set()  # Empty set since we allow email duplicates
        
        created_students = []
        errors = []
        
        for row in rows:
            student_name = str(row.get('Student Name', '')).strip()
            roll_number = str(row.get('Roll Number', '')).strip()
            email = str(row.get('Email', '')).strip().lower()
            mobile_number = str(row.get('Mobile Number', '')).strip()
            
            # Validation
            if not all([student_name, roll_number, email]):
                errors.append(f"{student_name or roll_number or email}: Missing required fields")
                continue
            
            if roll_number in existing_roll_numbers:
                errors.append(f"{student_name}: Roll number already exists")
                continue
            
            if email in existing_emails:
                errors.append(f"{student_name}: Email already exists")
                continue
            
            try:
                # Create user account
                username = roll_number
                password = f"{student_name.split()[0][:4].lower()}{roll_number[-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                
                user_doc = {
                    'username': username,
                    'email': email,
                    'password_hash': password_hash,
                    'role': 'student',
                    'name': student_name,
                    'mobile_number': mobile_number,
                    'campus_id': batch['campus_ids'][0] if batch.get('campus_ids') else None,
                    'course_id': course['_id'],
                    'batch_id': batch['_id'],
                    'batch_course_instance_id': ObjectId(instance_id),
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                
                user_id = mongo_db.users.insert_one(user_doc).inserted_id
                
                # Create default notification preferences for the new user
                try:
                    from models_notification_preferences import NotificationPreferences
                    NotificationPreferences.create_default_preferences(user_id)
                    current_app.logger.info(f"✅ Created default notification preferences for student: {user_id}")
                except Exception as e:
                    current_app.logger.warning(f"⚠️ Failed to create default notification preferences for student {user_id}: {e}")
                    # Don't fail user creation if notification preferences creation fails
                
                # Create student profile
                student_doc = {
                    'user_id': user_id,
                    'name': student_name,
                    'roll_number': roll_number,
                    'email': email,
                    'mobile_number': mobile_number,
                    'campus_id': batch['campus_ids'][0] if batch.get('campus_ids') else None,
                    'course_id': course['_id'],
                    'batch_id': batch['_id'],
                    'batch_course_instance_id': ObjectId(instance_id),
                    'created_at': datetime.now(pytz.utc)
                }
                
                mongo_db.students.insert_one(student_doc)
                
                created_students.append({
                    'name': student_name,
                    'email': email,
                    'username': username,
                    'password': password
                })
                
                # Update existing sets
                existing_roll_numbers.add(roll_number)
                existing_emails.add(email)
                
            except Exception as e:
                errors.append(f"{student_name}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Successfully created {len(created_students)} students',
            'data': {
                'created_students': created_students,
                'errors': errors
            }
        }), 201
        
    except Exception as e:
        current_app.logger.error(f"Error uploading students to instance: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/<batch_id>/verify-students', methods=['POST'])
@jwt_required()
def verify_students_upload(batch_id):
    """Verify that students were actually uploaded to the batch"""
    try:
        data = request.get_json()
        student_emails = data.get('student_emails', [])
        student_roll_numbers = data.get('student_roll_numbers', [])
        
        if not student_emails and not student_roll_numbers:
            return jsonify({'success': False, 'message': 'No student emails or roll numbers provided for verification.'}), 400
        
        batch_obj_id = ObjectId(batch_id)
        batch = mongo_db.batches.find_one({'_id': batch_obj_id})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        
        # Check for students in the batch
        query = {'batch_id': batch_obj_id}
        if student_emails:
            query['email'] = {'$in': student_emails}
        elif student_roll_numbers:
            query['roll_number'] = {'$in': student_roll_numbers}
        
        students = list(mongo_db.students.find(query))
        users = list(mongo_db.users.find({'batch_id': batch_obj_id, 'role': 'student'}))
        
        # Create lookup dictionaries
        student_lookup = {s['email']: s for s in students}
        user_lookup = {u['email']: u for u in users}
        
        verification_results = []
        for email in student_emails:
            student_exists = email in student_lookup
            user_exists = email in user_lookup
            
            verification_results.append({
                'email': email,
                'student_profile_exists': student_exists,
                'user_account_exists': user_exists,
                'fully_uploaded': student_exists and user_exists,
                'student_id': str(student_lookup[email]['_id']) if student_exists else None,
                'user_id': str(user_lookup[email]['_id']) if user_exists else None
            })
        
        total_students = len(student_emails)
        successful_uploads = sum(1 for r in verification_results if r['fully_uploaded'])
        failed_uploads = total_students - successful_uploads
        
        return jsonify({
            'success': True,
            'data': {
                'verification_results': verification_results,
                'summary': {
                    'total_students': total_students,
                    'successful_uploads': successful_uploads,
                    'failed_uploads': failed_uploads,
                    'success_rate': (successful_uploads / total_students * 100) if total_students > 0 else 0
                }
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error verifying students upload: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>/cleanup-failed-students', methods=['POST'])
@jwt_required()
def cleanup_failed_students(batch_id):
    """Clean up any orphaned user accounts or student profiles"""
    try:
        data = request.get_json()
        student_emails = data.get('student_emails', [])
        
        if not student_emails:
            return jsonify({'success': False, 'message': 'No student emails provided for cleanup.'}), 400
        
        batch_obj_id = ObjectId(batch_id)
        
        # Find orphaned records (users without corresponding student profiles or vice versa)
        cleanup_results = []
        
        for email in student_emails:
            user = mongo_db.users.find_one({'email': email, 'batch_id': batch_obj_id})
            student = mongo_db.students.find_one({'email': email, 'batch_id': batch_obj_id})
            
            if user and not student:
                # Orphaned user account - delete it
                mongo_db.users.delete_one({'_id': user['_id']})
                cleanup_results.append({
                    'email': email,
                    'action': 'deleted_orphaned_user',
                    'user_id': str(user['_id'])
                })
            elif student and not user:
                # Orphaned student profile - delete it
                mongo_db.students.delete_one({'_id': student['_id']})
                cleanup_results.append({
                    'email': email,
                    'action': 'deleted_orphaned_student',
                    'student_id': str(student['_id'])
                })
            elif not user and not student:
                cleanup_results.append({
                    'email': email,
                    'action': 'no_records_found'
                })
        
        return jsonify({
            'success': True,
            'message': f'Cleanup completed. {len(cleanup_results)} records processed.',
            'data': {
                'cleanup_results': cleanup_results
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error cleaning up failed students: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/students/<student_id>', methods=['DELETE'])
@jwt_required()
@require_permission(module='batch_management')
def delete_student_management(student_id):
    """Delete a student"""
    try:
        # Delete user account
        mongo_db.users.delete_one({'_id': ObjectId(student_id)})
        
        # Delete student profile
        mongo_db.students.delete_one({'user_id': ObjectId(student_id)})
        
        return jsonify({'success': True, 'message': 'Student deleted successfully'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error deleting student: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/bulk-delete', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def bulk_delete_students():
    """Bulk delete students"""
    try:
        data = request.get_json()
        student_ids = data.get('student_ids', [])
        
        if not student_ids:
            return jsonify({'success': False, 'message': 'No students selected for deletion'}), 400
        
        if not isinstance(student_ids, list):
            return jsonify({'success': False, 'message': 'Invalid student IDs format'}), 400
        
        # Validate that all IDs are valid ObjectIds
        valid_ids = []
        invalid_ids = []
        
        for student_id in student_ids:
            try:
                valid_ids.append(ObjectId(student_id))
            except Exception:
                invalid_ids.append(student_id)
        
        if invalid_ids:
            return jsonify({
                'success': False, 
                'message': f'Invalid student IDs: {", ".join(invalid_ids)}'
            }), 400
        
        # Check if students exist
        existing_students = list(mongo_db.users.find(
            {'_id': {'$in': valid_ids}}, 
            {'_id': 1, 'name': 1}
        ))
        
        existing_ids = [str(student['_id']) for student in existing_students]
        missing_ids = [str(sid) for sid in valid_ids if str(sid) not in existing_ids]
        
        if missing_ids:
            return jsonify({
                'success': False,
                'message': f'Students not found: {", ".join(missing_ids)}'
            }), 404
        
        # Perform bulk deletion
        deleted_count = 0
        failed_deletions = []
        
        for student_id in valid_ids:
            try:
                # Delete user account
                user_result = mongo_db.users.delete_one({'_id': student_id})
                
                # Delete student profile
                student_result = mongo_db.students.delete_one({'user_id': student_id})
                
                if user_result.deleted_count > 0 or student_result.deleted_count > 0:
                    deleted_count += 1
                else:
                    failed_deletions.append(str(student_id))
                    
            except Exception as e:
                current_app.logger.error(f"Error deleting student {student_id}: {e}")
                failed_deletions.append(str(student_id))
        
        # Prepare response
        response_data = {
            'success': True,
            'deleted_count': deleted_count,
            'total_requested': len(valid_ids)
        }
        
        if failed_deletions:
            response_data['failed_deletions'] = failed_deletions
            response_data['message'] = f'Successfully deleted {deleted_count} students. Failed to delete {len(failed_deletions)} students.'
        else:
            response_data['message'] = f'Successfully deleted {deleted_count} students.'
        
        return jsonify(response_data), 200
        
    except Exception as e:
        current_app.logger.error(f"Error in bulk delete students: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/students/<student_id>/send-credentials', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def send_student_credentials(student_id):
    """Send credentials to student"""
    try:
        student = mongo_db.users.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        # Get student profile for roll number
        student_profile = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        
        # Generate password using consistent pattern: first 4 letters of name + last 4 digits of roll number
        student_name = student.get('name', '')
        roll_number = student_profile.get('roll_number', '') if student_profile else ''
        
        if student_name and roll_number:
            # Extract first 4 letters of first name and last 4 digits of roll number
            first_name = student_name.split()[0] if student_name.split() else student_name
            password = f"{first_name[:4].lower()}{roll_number[-4:]}"
        else:
            # Fallback to random password if name or roll number is missing
            import secrets
            import string
            password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
        
        # Update password
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        mongo_db.users.update_one(
            {'_id': ObjectId(student_id)},
            {'$set': {'password_hash': password_hash}}
        )
        
        # Send email with credentials
        subject = "Your Study Edge Login Credentials"
        html_content = render_template(
            'student_credentials.html',
            params={
                'name': student.get('name', ''),
                'email': student.get('email', ''),
                'username': student.get('username', ''),
                'password': password,
                'roll_number': student_profile.get('roll_number', '') if student_profile else '',
                'login_url': "https://crt.pydahsoft.in/login"
            }
        )
        
        send_email(
            to_email=student.get('email'),
            to_name=student.get('name', 'Student'),
            subject=subject,
            html_content=html_content
        )
        
        return jsonify({'success': True, 'message': 'Credentials sent successfully'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error sending credentials: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/bulk-migrate-progress', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def bulk_migrate_students_progress():
    """Bulk migrate all students to new progress system"""
    try:
        current_user_id = get_jwt_identity()
        current_user = mongo_db.users.find_one({'_id': ObjectId(current_user_id)})
        
        if not current_user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Debug: Log user data to see what fields are available
        current_app.logger.info(f"User data for migration: {current_user}")
        
        # Check permissions - only superadmin can do bulk migration
        allowed_roles = ['superadmin', 'super_admin','sub_superadmin']
        if current_user.get('role') not in allowed_roles:
            return jsonify({'success': False, 'message': 'Only superadmin can perform bulk migration'}), 403
        
        # Get user display name (try multiple fields)
        user_name = (current_user.get('name') or 
                    current_user.get('username') or 
                    current_user.get('email') or 
                    current_user.get('full_name') or 
                    f"User {current_user_id}")
        
        current_app.logger.info(f"Starting bulk migration by user {user_name} (ID: {current_user_id}, Role: {current_user.get('role', 'Unknown')})")
        
        # Find all students that need migration (have old string-based authorized_levels)
        students_to_migrate = list(mongo_db.students.find({
            'authorized_levels': {
                '$exists': True,
                '$ne': [],
                '$elemMatch': {'$type': 'string'}  # Contains at least one string element
            }
        }))
        
        if not students_to_migrate:
            return jsonify({
                'success': True,
                'message': 'No students need migration - all are already using the new progress system',
                'migrated_count': 0,
                'failed_count': 0,
                'already_migrated_count': 0,
                'total_students_processed': 0,
                'migration_errors': []
            }), 200
        
        # Also count already migrated students
        already_migrated_count = mongo_db.students.count_documents({
            'authorized_levels': {
                '$exists': True,
                '$ne': [],
                '$elemMatch': {'$type': 'object'}  # Contains at least one object element
            }
        })
        
        migrated_count = 0
        failed_count = 0
        migration_errors = []
        
        try:
            from utils.student_progress_manager import StudentProgressManager
            # PASS mongo_db to StudentProgressManager - previously missing which caused runtime errors
            progress_manager = StudentProgressManager(mongo_db)
            
            for student in students_to_migrate:
                try:
                    authorized_levels = student.get('authorized_levels', [])

                    # Preserve existing object entries and convert only string entries
                    existing_objects = [lvl for lvl in authorized_levels if isinstance(lvl, dict)]
                    string_entries = [lvl for lvl in authorized_levels if isinstance(lvl, str)]

                    converted_from_strings = []
                    for level_id in string_entries:
                        converted_from_strings.append({
                            'level_id': level_id,
                            'unlocked_at': datetime.now(),
                            'unlock_source': 'legacy_migration',
                            'unlocked_by': current_user_id,
                            'reason': 'Bulk migrated from old progress system'
                        })

                    # Merge existing object entries with converted entries, prefer existing object metadata
                    merged_map = {obj.get('level_id'): obj for obj in existing_objects if obj.get('level_id')}
                    for conv in converted_from_strings:
                        lid = conv.get('level_id')
                        if lid and lid not in merged_map:
                            merged_map[lid] = conv

                    new_authorized_levels = list(merged_map.values())

                    # Prepare update_data while preserving/initializing other fields
                    update_data = {
                        'authorized_levels': new_authorized_levels,
                        'migration_status': 'completed',
                        'migrated_at': datetime.now(),
                        'migrated_by': current_user_id
                    }

                    # Ensure module_progress exists (do not overwrite if present)
                    if 'module_progress' not in student:
                        update_data['module_progress'] = {}

                    # Ensure unlock_history exists - append converted history entries if unlock_history exists
                    existing_history = student.get('unlock_history', []) or []
                    converted_history_entries = []
                    for level_id in string_entries:
                        converted_history_entries.append({
                            'level_id': level_id,
                            'unlocked_at': datetime.now(),
                            'unlocked_by': 'legacy_migration',
                            'score': None,
                            'test_id': None
                        })

                    # Merge histories (append converted entries while avoiding duplicates by level_id+timestamp)
                    # Simple append is fine for now; dedupe by level_id
                    existing_history_map = {h.get('level_id'): h for h in existing_history if isinstance(h, dict) and h.get('level_id')}
                    for h in converted_history_entries:
                        if h.get('level_id') not in existing_history_map:
                            existing_history.append(h)

                    update_data['unlock_history'] = existing_history

                    result = mongo_db.students.update_one(
                        {'_id': student['_id']},
                        {'$set': update_data}
                    )
                    
                    if result.modified_count > 0:
                        migrated_count += 1
                        
                        # Log the migration event
                        progress_manager.log_progress_event(
                            event_type='student_migration',
                            student_id=str(student['_id']),
                            level_id=None,
                            details={
                                'migrated_by': current_user_id,
                                'old_levels_count': len(authorized_levels),
                                'new_levels_count': len(new_authorized_levels),
                                'migration_reason': 'Bulk migration via admin interface',
                                'student_name': student.get('name', 'Unknown'),
                                'roll_number': student.get('roll_number', 'N/A')
                            }
                        )
                        
                        current_app.logger.info(f"Migrated student: {student.get('name', 'Unknown')} ({student.get('roll_number', 'N/A')}) by {user_name}")
                    else:
                        failed_count += 1
                        migration_errors.append(f"Failed to update student {student.get('name', 'Unknown')}")
                        
                except Exception as e:
                    failed_count += 1
                    error_msg = f"Error migrating student {student.get('name', 'Unknown')}: {str(e)}"
                    migration_errors.append(error_msg)
                    current_app.logger.error(error_msg)
            
            # Log bulk migration completion
            progress_manager.log_progress_event(
                event_type='bulk_migration_completed',
                student_id=None,
                level_id=None,
                details={
                    'migrated_by': current_user_id,
                    'total_migrated': migrated_count,
                    'total_failed': failed_count,
                    'already_migrated': already_migrated_count,
                    'migration_timestamp': datetime.now().isoformat()
                }
            )
            
            current_app.logger.info(f"Bulk migration completed by {user_name}: {migrated_count} migrated, {failed_count} failed, {already_migrated_count} already migrated")
            
            return jsonify({
                'success': True,
                'message': f'Bulk migration completed successfully',
                'migrated_count': migrated_count,
                'failed_count': failed_count,
                'already_migrated_count': already_migrated_count,
                'total_students_processed': len(students_to_migrate),
                'migration_errors': migration_errors[:10],  # Limit to first 10 errors
                'migrated_at': datetime.now().isoformat()
            }), 200
                
        except Exception as e:
            current_app.logger.error(f"Error in bulk migration process: {e}")
            return jsonify({'success': False, 'message': f'Bulk migration failed: {str(e)}'}), 500
        
    except Exception as e:
        current_app.logger.error(f"Error in bulk_migrate_students_progress: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/<student_id>/credentials', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def download_student_credentials(student_id):
    """Download student credentials as CSV"""
    try:
        student = mongo_db.users.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        student_profile = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        
        # Generate password using consistent pattern: first 4 letters of name + last 4 digits of roll number
        student_name = student.get('name', '')
        roll_number = student_profile.get('roll_number', '') if student_profile else ''
        
        if student_name and roll_number:
            # Extract first 4 letters of first name and last 4 digits of roll number
            first_name = student_name.split()[0] if student_name.split() else student_name
            password = f"{first_name[:4].lower()}{roll_number[-4:]}"
        else:
            # Fallback to random password if name or roll number is missing
            import secrets
            import string
            password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
        
        # Update password
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        mongo_db.users.update_one(
            {'_id': ObjectId(student_id)},
            {'$set': {'password_hash': password_hash}}
        )
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Email', 'Roll Number', 'Username', 'Password'])
        writer.writerow([
            student.get('name', ''),
            student.get('email', ''),
            student_profile.get('roll_number', '') if student_profile else '',
            student.get('username', ''),
            password
        ])
        
        output.seek(0)
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{student.get("name", "student")}_credentials.csv"'
        }
        
    except Exception as e:
        current_app.logger.error(f"Error downloading credentials: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/filtered/credentials', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def download_all_filtered_students_credentials():
    """Download credentials for all filtered students as CSV"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Get query parameters (same as filtered students endpoint)
        search = request.args.get('search', '')
        campus_id = request.args.get('campus_id', '')
        course_id = request.args.get('course_id', '')
        batch_id = request.args.get('batch_id', '')
        
        # Build query (same as filtered students endpoint)
        query = {'role': 'student'}
        
        if search:
            query['$or'] = [
                {'name': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}}
            ]
        
        if campus_id:
            query['campus_id'] = ObjectId(campus_id)
        
        if course_id:
            query['course_id'] = ObjectId(course_id)
        
        if batch_id:
            query['batch_id'] = ObjectId(batch_id)
        
        # Super admin and sub_superadmin can see all students, others only their campus
        if user.get('role') not in ['superadmin', 'sub_superadmin']:
            user_campus_id = user.get('campus_id')
            if user_campus_id:
                query['campus_id'] = ObjectId(user_campus_id)
        
        # Get all students matching the filter (no pagination)
        students = list(mongo_db.users.find(query))
        
        if not students:
            return jsonify({'success': False, 'message': 'No students found matching the criteria'}), 404
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Email', 'Roll Number', 'Username', 'Password', 'Campus', 'Course', 'Batch'])
        
        for student in students:
            # Get student profile
            student_profile = mongo_db.students.find_one({'user_id': student['_id']})
            
            # Get campus, course, and batch names
            campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')})
            course = mongo_db.courses.find_one({'_id': student.get('course_id')})
            batch = mongo_db.batches.find_one({'_id': student.get('batch_id')})
            
            # Generate password using consistent pattern
            student_name = student.get('name', '')
            roll_number = student_profile.get('roll_number', '') if student_profile else ''
            
            if student_name and roll_number:
                first_name = student_name.split()[0] if student_name.split() else student_name
                password = f"{first_name[:4].lower()}{roll_number[-4:]}"
            else:
                import secrets
                import string
                password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
            
            # Update password in database
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            mongo_db.users.update_one(
                {'_id': student['_id']},
                {'$set': {'password_hash': password_hash}}
            )
            
            writer.writerow([
                student.get('name', ''),
                student.get('email', ''),
                student_profile.get('roll_number', '') if student_profile else '',
                student.get('username', ''),
                password,
                campus.get('name', '') if campus else '',
                course.get('name', '') if course else '',
                batch.get('name', '') if batch else ''
            ])
        
        output.seek(0)
        
        # Generate filename based on filters
        filename_parts = ['students_credentials']
        if campus_id:
            campus = mongo_db.campuses.find_one({'_id': ObjectId(campus_id)})
            if campus:
                filename_parts.append(campus.get('name', '').replace(' ', '_'))
        if course_id:
            course = mongo_db.courses.find_one({'_id': ObjectId(course_id)})
            if course:
                filename_parts.append(course.get('name', '').replace(' ', '_'))
        if batch_id:
            batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
            if batch:
                filename_parts.append(batch.get('name', '').replace(' ', '_'))
        if search:
            filename_parts.append('search')
        
        filename = '_'.join(filename_parts) + '.csv'
        
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
    except Exception as e:
        current_app.logger.error(f"Error downloading all credentials: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/batch/<batch_id>/send-emails', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def send_batch_emails(batch_id):
    """Send welcome emails to all students in a batch."""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Verify batch exists and user has access
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found'}), 404

        # Get all students in the batch
        students = list(mongo_db.students.find({'batch_id': ObjectId(batch_id)}))
        if not students:
            return jsonify({'success': False, 'message': 'No students found in this batch'}), 404

        # Get user details for each student
        student_details = []
        for student in students:
            user_detail = mongo_db.users.find_one({'_id': student['user_id']})
            if user_detail:
                # Generate password if not present (same logic as manual_student_notifications.py)
                existing_password = user_detail.get('password', '')
                if not existing_password:
                    # Generate password using the same pattern as during upload
                    first_name = student['name'].split()[0] if student['name'].split() else student['name']
                    roll_number = student.get('roll_number', '')
                    if first_name and roll_number:
                        generated_password = f"{first_name[:4].lower()}{roll_number[-4:]}"
                    else:
                        # Fallback to random password
                        import secrets
                        import string
                        generated_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
                    
                    # Update the user record with the generated password
                    mongo_db.users.update_one(
                        {'_id': student['user_id']},
                        {'$set': {'password': generated_password}}
                    )
                    password = generated_password
                else:
                    password = existing_password
                
                student_details.append({
                    'student_id': str(student['_id']),
                    'name': student['name'],
                    'email': student.get('email'),
                    'username': user_detail.get('username', ''),
                    'password': password
                })

        # Filter students with email addresses
        students_with_email = [s for s in student_details if s['email']]
        total_students = len(students_with_email)

        if total_students == 0:
            return jsonify({'success': False, 'message': 'No students with email addresses found in this batch'}), 400

        # Prepare students for batch processing
        students_for_batch = []
        for student in students_with_email:
            student_data = {
                'student_id': student['student_id'],
                'name': student['name'],
                'email': student['email'],
                'mobile_number': student.get('mobile_number'),
                'username': student['username'],
                'password': student['password']
            }
            students_for_batch.append(student_data)
        
        # Create batch job for email-only processing
        try:
            from utils.batch_processor import create_email_only_batch_job
            
            batch_result = create_email_only_batch_job(
                students=students_for_batch,
                batch_size=100,
                interval_minutes=3
            )
            
            current_app.logger.info(f"📧📱 Email batch job created: {batch_result}")
            
            # Send progress update
            socketio.emit('email_progress', {
                'user_id': current_user_id,
                'status': 'queued',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
                'message': f'Email batch job created successfully! {len(students_for_batch)} emails queued for background processing',
                'batch_id': batch_result.get('batch_id'),
                'estimated_completion': batch_result.get('estimated_completion')
        }, room=str(current_user_id))
            
        except Exception as e:
            current_app.logger.error(f"❌ Failed to create email batch job: {e}")
            # Fallback to individual email processing
            try:
                from utils.notification_queue import queue_batch_notifications
                # Queue notifications individually as a fallback
                queue_result = queue_batch_notifications(students_for_batch, notification_type='welcome')
                current_app.logger.info(f"Fallback queued individual emails: {queue_result}")
            except Exception as fallback_err:
                current_app.logger.error(f"Fallback email processing also failed: {fallback_err}")

        return jsonify({
            'success': True,
            'message': f'Email batch job created successfully. {len(students_for_batch)} emails queued for background processing.',
            'data': {
                'total_students': total_students,
                'emails_queued': len(students_for_batch),
                'batch_id': batch_result.get('batch_id') if 'batch_result' in locals() else None,
                'estimated_completion': batch_result.get('estimated_completion') if 'batch_result' in locals() else None
            }
        }), 200

    except Exception as e:
        current_app.logger.error(f"Error sending batch emails: {e}")
        return jsonify({'success': False, 'message': f'Failed to send emails: {str(e)}'}), 500

@batch_management_bp.route('/batch/<batch_id>/send-sms', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def send_batch_sms(batch_id):
    """Send SMS notifications to all students in a batch."""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Verify batch exists and user has access
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found'}), 404

        # Get all students in the batch
        students = list(mongo_db.students.find({'batch_id': ObjectId(batch_id)}))
        if not students:
            return jsonify({'success': False, 'message': 'No students found in this batch'}), 404

        # Get user details for each student
        student_details = []
        for student in students:
            user_detail = mongo_db.users.find_one({'_id': student['user_id']})
            if user_detail:
                # Generate password if not present (same logic as manual_student_notifications.py)
                existing_password = user_detail.get('password', '')
                if not existing_password:
                    # Generate password using the same pattern as during upload
                    first_name = student['name'].split()[0] if student['name'].split() else student['name']
                    roll_number = student.get('roll_number', '')
                    if first_name and roll_number:
                        generated_password = f"{first_name[:4].lower()}{roll_number[-4:]}"
                    else:
                        # Fallback to random password
                        import secrets
                        import string
                        generated_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
                    
                    # Update the user record with the generated password
                    mongo_db.users.update_one(
                        {'_id': student['user_id']},
                        {'$set': {'password': generated_password}}
                    )
                    password = generated_password
                else:
                    password = existing_password
                
                student_details.append({
                    'student_id': str(student['_id']),
                    'name': student['name'],
                    'mobile_number': student.get('mobile_number'),
                    'username': user_detail.get('username', ''),
                    'password': password
                })

        # Filter students with mobile numbers
        students_with_mobile = [s for s in student_details if s['mobile_number']]
        total_students = len(students_with_mobile)

        if total_students == 0:
            return jsonify({'success': False, 'message': 'No students with mobile numbers found in this batch'}), 400

        # Create resilient services
        resilient_services = create_resilient_services()
        
        # Track results
        sms_results = []
        successful_sms = 0
        failed_sms = 0

        # Send SMS with progress tracking
        for index, student in enumerate(students_with_mobile):
            student_name = student['name']
            mobile_number = student['mobile_number']
            username = student['username']
            password = student['password']
            
            result = {
                'student_id': student['student_id'],
                'name': student_name,
                'mobile_number': mobile_number,
                'sms_sent': False,
                'error': None
            }
            
            try:
                # Send SMS using resilient service
                sms_result = resilient_services['sms'].send_sms_resilient(
                    phone=mobile_number,
                    student_name=student_name,
                    username=username,
                    password=password,
                    login_url="https://crt.pydahsoft.in/login"
                )
                
                result['sms_sent'] = sms_result.get('success', False)
                if sms_result.get('success'):
                    successful_sms += 1
                    current_app.logger.info(f"✅ SMS sent to {mobile_number}")
                else:
                    failed_sms += 1
                    result['error'] = sms_result.get('error', 'SMS sending failed after retries')
                    current_app.logger.warning(f"⚠️ SMS failed for {mobile_number}: {result['error']}")
                    
            except Exception as e:
                failed_sms += 1
                result['error'] = str(e)
                current_app.logger.error(f"❌ SMS error for {mobile_number}: {e}")
            
            sms_results.append(result)
            
            # Send progress update
            percentage = int(((index + 1) / total_students) * 100)
            socketio.emit('sms_progress', {
                'user_id': current_user_id,
                'status': 'sending_sms',
                'total': total_students,
                'processed': index + 1,
                'percentage': percentage,
                'message': f'Sending SMS: {student_name} - {"✅" if result["sms_sent"] else "❌"}',
                'current_student': {
                    'name': student_name,
                    'mobile_number': mobile_number,
                    'sms_sent': result['sms_sent']
                }
            }, room=str(current_user_id))

        # Send completion notification
        socketio.emit('sms_progress', {
            'user_id': current_user_id,
            'status': 'completed',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
            'message': f'SMS sending completed! {successful_sms} sent, {failed_sms} failed'
        }, room=str(current_user_id))

        return jsonify({
            'success': True,
            'message': f'SMS sending completed. {successful_sms} SMS sent, {failed_sms} failed.',
            'data': {
                'total_students': total_students,
                'successful_sms': successful_sms,
                'failed_sms': failed_sms,
                'sms_results': sms_results
            }
        }), 200

    except Exception as e:
        current_app.logger.error(f"Error sending batch SMS: {e}")
        return jsonify({'success': False, 'message': f'Failed to send SMS: {str(e)}'}), 500

@batch_management_bp.route('/notification-stats', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_notification_stats_endpoint():
    """Get notification queue statistics"""
    try:
        stats = get_notification_stats()
        return jsonify({
            'success': True,
            'data': stats
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error getting notification stats: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500